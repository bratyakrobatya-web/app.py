import streamlit as st
import requests
import pandas as pd
import numpy as np
from rapidfuzz import fuzz, process
import io
import re
import zipfile
from datetime import datetime
import os
from typing import Dict, List, Optional

# Security utilities
from security_utils import (
    RateLimiter,
    logger,
    log_security_event,
    sanitize_html,
    sanitize_user_input,
    sanitize_csv_content,
    validate_file_size,
    validate_file_extension,
    safe_session_append,
    cleanup_session_state,
    check_session_state_limits,
    handle_production_error,
    MAX_FILE_SIZE,
    MAX_FILES_COUNT,
    ALLOWED_FILE_EXTENSIONS
)
from requests.exceptions import RequestException, Timeout, HTTPError

# Safe file operations
from safe_file_utils import (
    safe_open_image,
    safe_read_csv,
    safe_read_file
)

# City matching module
from modules.matching import (
    normalize_city_name,
    extract_city_and_region,
    get_candidates_by_word,
    PREFERRED_MATCHES,
    EXCLUDED_EXACT_MATCHES
)

# Data processing module
from modules.data_processing import (
    get_hh_areas,
    load_population_data,
    get_federal_district_by_region,
    get_cities_by_regions,
    get_all_cities,
    normalize_region_name,
    FEDERAL_DISTRICTS
)

# Utility functions module
from modules.utils import (
    get_russian_cities,
    remove_header_row_if_needed,
    check_if_changed
)

# City matcher module
from modules.city_matcher import (
    smart_match_city,
    match_cities,
    merge_cities_files
)

# Export utilities module
from modules.export_utils import (
    create_excel_buffer,
    create_publisher_excel,
    create_full_report_excel,
    create_zip_archive,
    create_result_excel
)

# ============================================
# PERFORMANCE OPTIMIZATION: Cached Functions
# ============================================

@st.cache_data(show_spinner=False)
def get_russian_cities_cached(_hh_areas: Dict) -> List[str]:
    """
    Кэшированная версия get_russian_cities для оптимизации производительности.

    Использует st.cache_data для кэширования результата, чтобы избежать повторной
    фильтрации ~18,000 городов при каждом rerun.

    Args:
        _hh_areas: Справочник регионов HH.ru (префикс _ для bypass hashing)

    Returns:
        List[str]: Список названий городов России
    """
    return get_russian_cities(_hh_areas)


@st.cache_data(show_spinner=False)
def prepare_city_options(candidates: tuple, current_value: str, current_match: float, city_name: str) -> tuple:
    """
    Кэшированная подготовка options для selectbox.

    ОПТИМИЗАЦИЯ: избегаем пересоздания списков и сортировки на каждом rerun.
    Используется tuple для candidates чтобы можно было кэшировать.

    Args:
        candidates: Кортеж кандидатов (city_name, match_percent)
        current_value: Текущее значение гео
        current_match: Процент совпадения текущего значения
        city_name: Исходное название города

    Returns:
        tuple: (options, candidates_dict)
            - options: список для selectbox
            - candidates_dict: {city_name: index} для O(1) поиска
    """
    candidates_list = list(candidates)

    # Добавляем текущее значение если его нет
    if current_value and current_value != city_name:
        candidate_names = [c[0] for c in candidates_list]
        if current_value not in candidate_names:
            candidates_list.append((current_value, current_match))

    # Сортируем по убыванию процента
    candidates_list.sort(key=lambda x: x[1], reverse=True)

    # Формируем options
    if candidates_list:
        options = ["❌ Нет совпадения"] + [f"{c[0]} ({c[1]:.1f}%)" for c in candidates_list[:20]]
    else:
        options = ["❌ Нет совпадения"]

    # Создаём словарь для O(1) поиска индекса по названию города
    candidates_dict = {c[0]: i + 1 for i, c in enumerate(candidates_list[:20])}

    return tuple(options), candidates_dict


@st.cache_data(show_spinner="Загрузка справочника HH.ru...", ttl=3600)
def get_hh_areas_cached() -> Optional[Dict]:
    """
    Кэшированная версия get_hh_areas для критической оптимизации производительности.

    БЕЗ кэширования:
    - HTTP запрос к API при КАЖДОМ rerun (~300-500ms)
    - Парсинг JSON с 18,000 городами при КАЖДОМ изменении виджета (~200-300ms)
    - Полная перерисовка страницы занимает 500-800ms

    С кэшированием:
    - Запрос выполняется 1 раз в час
    - Все последующие reruns используют кэшированный результат
    - Время rerun сокращается до ~50-100ms

    Args:
        None

    Returns:
        Optional[Dict]: Справочник регионов HH.ru или None при ошибке
    """
    return get_hh_areas()


@st.cache_data(show_spinner=False)
def apply_manual_selections_cached(_result_df, manual_selections: dict, _hh_areas: dict, cache_key: str = "default") -> pd.DataFrame:
    """
    Кэшированное применение ручных изменений к DataFrame.

    КРИТИЧНО ДЛЯ ПРОИЗВОДИТЕЛЬНОСТИ:
    - БЕЗ кэша: применяется при КАЖДОМ rerun (~1000ms для 30 городов)
    - С кэшем: применяется ТОЛЬКО при изменении manual_selections (~5ms)

    ВАЖНО: manual_selections БЕЗ _ чтобы Streamlit хэшировал СОДЕРЖИМОЕ словаря!
    При изменении значений внутри словаря хэш меняется → кэш инвалидируется → функция выполняется.

    FIX: Добавлен cache_key для различения разных вкладок/вакансий.
    Без этого параметра все вкладки с пустым manual_selections={} получали одинаковый кэшированный результат!

    Args:
        _result_df: Исходный DataFrame с результатами (НЕ хэшируется)
        manual_selections: Словарь ручных изменений {row_id: new_value} (ХЭШИРУЕТСЯ!)
        _hh_areas: Справочник HH.ru (НЕ хэшируется)
        cache_key: Уникальный ключ для кэша (название вкладки/вакансии) (ХЭШИРУЕТСЯ!)

    Returns:
        pd.DataFrame: DataFrame с применёнными изменениями
    """
    # Копируем только если есть изменения
    if not manual_selections:
        return _result_df

    final_df = _result_df.copy()

    # Применяем ручные изменения
    for row_id, new_value in manual_selections.items():
        mask = final_df['row_id'] == row_id

        # FIX: Проверяем что row_id существует в DataFrame
        if mask.sum() == 0:
            continue  # Пропускаем если строка не найдена

        if new_value == "❌ Нет совпадения":
            final_df.loc[mask, 'Итоговое гео'] = None
            final_df.loc[mask, 'ID HH'] = None
            final_df.loc[mask, 'Регион'] = None
            final_df.loc[mask, 'Совпадение %'] = 0
            final_df.loc[mask, 'Изменение'] = 'Нет'
            final_df.loc[mask, 'Статус'] = '❌ Не найдено'
        else:
            final_df.loc[mask, 'Итоговое гео'] = new_value

            if new_value in _hh_areas:
                final_df.loc[mask, 'ID HH'] = _hh_areas[new_value]['id']
                final_df.loc[mask, 'Регион'] = _hh_areas[new_value]['parent']

            original = final_df.loc[mask, 'Исходное название'].values[0]
            final_df.loc[mask, 'Изменение'] = 'Да' if check_if_changed(original, new_value) else 'Нет'

    return final_df

# Version: 3.3.2 - Fixed: corrected all indentation in single mode block

@st.cache_data(show_spinner=False)
def get_cached_icon_base64(filename: str) -> Optional[str]:
    """
    Кэшированная загрузка иконки и конвертация в base64.

    ОПТИМИЗАЦИЯ: иконки загружаются 1 раз при старте вместо каждого rerun.
    Экономия: ~10-20ms на каждую иконку при каждом rerun.

    Args:
        filename: имя файла иконки

    Returns:
        base64 строка или None
    """
    from io import BytesIO
    import base64

    icon_image = safe_open_image(filename)
    if icon_image:
        buffered = BytesIO()
        icon_image.save(buffered, format="PNG")
        img_base64 = base64.b64encode(buffered.getvalue()).decode()
        return f"data:image/png;base64,{img_base64}"
    return None


@st.cache_data(show_spinner=False)
def create_excel_bytes_cached(df: pd.DataFrame, sheet_name: str) -> bytes:
    """
    Кэшированная генерация Excel файла.
    
    КРИТИЧНО ДЛЯ ПРОИЗВОДИТЕЛЬНОСТИ:
    - Избегает пересоздания Excel файла при каждом rerun, если данные не изменились.
    - Решает проблему "подвисания" при переключении вкладок и выборе городов.
    """
    # Санитизация данных перед экспортом (защита от CSV Injection)
    safe_df = sanitize_csv_content(df)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        safe_df.to_excel(writer, index=False, header=True, sheet_name='Результат')
    return buffer.getvalue()


@st.cache_data(show_spinner=False)
def prepare_final_sheet_output_cached(result_df: pd.DataFrame, original_df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    Кэшированная подготовка итогового DataFrame для вкладки.
    
    ОПТИМИЗАЦИЯ:
    - Выполняет фильтрацию, merge и очистку данных только один раз.
    - Предотвращает повторные тяжелые вычисления при перерисовке интерфейса.
    """
    # 1. Фильтрация валидных строк
    output_df = result_df[
        (result_df['Итоговое гео'].notna()) &
        (~result_df['Статус'].str.contains('❌ Не найдено', na=False)) &
        (~result_df['Статус'].str.contains('Пустое значение', na=False))
    ].copy()

    # 2. Исключение дубликатов городов с "❌ Не найдено"
    excluded_cities = result_df[
        result_df['Статус'].str.contains('❌ Не найдено', na=False)
    ]['Исходное название'].unique()

    if len(excluded_cities) > 0:
        excluded_normalized = set()
        for city in excluded_cities:
            if pd.notna(city):
                normalized = str(city).replace('ё', 'е').replace('Ё', 'Е').lower().strip()
                normalized = ' '.join(normalized.split())
                excluded_normalized.add(normalized)

        output_df['_temp_normalized'] = (
            output_df['Исходное название']
            .fillna('').astype(str)
            .str.replace('ё', 'е').str.replace('Ё', 'Е')
            .str.lower().str.strip()
            .str.replace(r'\s+', ' ', regex=True)
        )
        output_df = output_df[~output_df['_temp_normalized'].isin(excluded_normalized)].copy()
        output_df = output_df.drop(columns=['_temp_normalized'])

    if len(output_df) == 0:
        return pd.DataFrame()

    # 3. Объединение с исходными данными
    original_cols = original_df.columns.tolist()
    final_output = pd.DataFrame()
    final_output[original_cols[0]] = output_df['Итоговое гео']

    for col in original_cols[1:]:
        if col in original_df.columns:
            temp_df = original_df.reset_index()
            temp_df['row_id'] = temp_df.index
            merged = output_df[['row_id']].merge(
                temp_df[['row_id', col]],
                on='row_id',
                how='left'
            )
            final_output[col] = merged[col].values
    
    # 4. Удаление дубликатов
    final_output['_normalized'] = (
        final_output[original_cols[0]]
        .fillna('').astype(str)
        .str.replace('ё', 'е').str.replace('Ё', 'Е')
        .str.lower().str.strip()
        .str.replace(r'\s+', ' ', regex=True)
    )
    final_output = final_output.drop_duplicates(subset=['_normalized'], keep='first')
    final_output = final_output.drop(columns=['_normalized'])

    # 5. Удаление заголовка если нужно
    final_output = remove_header_row_if_needed(final_output, original_cols[0])
    
    return final_output


# ============================================
# КОНФИГУРАЦИЯ: API КЛЮЧИ
# ============================================
# Для безопасного хранения ключа используется следующий приоритет:
# 1. Streamlit secrets (.streamlit/secrets.toml)
# 2. Переменная окружения ANTHROPIC_API_KEY
#
# Для настройки создайте файл .streamlit/secrets.toml с содержимым:
# ANTHROPIC_API_KEY = "ваш-ключ-здесь"

# Настройка страницы
st.set_page_config(
    page_title="Синхронизатор",
    page_icon="🌍",
    layout="wide"
)

# Кастомный CSS для современного дизайна
# Безопасная загрузка CSS из отдельного файла
css_content = safe_read_file("static/styles.css")
if css_content:
    st.markdown(f"<style>{css_content}</style>", unsafe_allow_html=True)
else:
    logger.error("Не удалось загрузить static/styles.css, стили не применены")


# Инициализация session_state
if 'result_df' not in st.session_state:
    st.session_state.result_df = None
if 'duplicate_count' not in st.session_state:
    st.session_state.duplicate_count = 0
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'manual_selections' not in st.session_state:
    st.session_state.manual_selections = {}
if 'candidates_cache' not in st.session_state:
    st.session_state.candidates_cache = {}
if 'search_query' not in st.session_state:
    st.session_state.search_query = ""
if 'added_cities' not in st.session_state:
    st.session_state.added_cities = []
if 'original_df' not in st.session_state:
    st.session_state.original_df = None

# Проверка лимитов session_state для предотвращения утечек памяти
session_stats = check_session_state_limits()
if session_stats['warnings']:
    for warning in session_stats['warnings']:
        logger.warning(f"Session state warning: {warning}")


# ============================================
# ИНТЕРФЕЙС
# ============================================

# Загрузка иконки synchronize.png с защитой от Path Traversal
# OPTIMIZED: use cached icon loading
try:
    sync_icon_base64 = get_cached_icon_base64("synchronize.png")
    if sync_icon_base64:
        SYNC_ICON = f'<img src="{sync_icon_base64}" style="width: 1em; height: 1em; display: inline-block;">'
    else:
        logger.warning("Не удалось загрузить synchronize.png, используется эмодзи")
        SYNC_ICON = '🔄'
except Exception as e:
    # Fallback если файл не найден
    SYNC_ICON = '🔄'

# Загрузка справочника HH с кэшированием (критично для производительности!)
# Без кэширования: HTTP запрос при КАЖДОМ rerun = 500-800ms задержка
# С кэшированием: запрос 1 раз в час, остальное из кэша = ~50ms
hh_areas = get_hh_areas_cached()

# ============================================
# ГЛАВНЫЙ ЗАГОЛОВОК
# ============================================
st.markdown('''
<div style="margin-bottom: 2rem;">
    <h1 style="text-align: left; color: #f4301f; margin-bottom: 0.3rem;">Синхронизатор</h1>
</div>
''', unsafe_allow_html=True)
st.markdown("---")

# ============================================
# БЛОК: ПРОВЕРКА ГЕО
# ============================================
if hh_areas:
    st.markdown('<div id="проверка-гео"></div>', unsafe_allow_html=True)
    st.header("🔍 Проверка гео и выгрузка базы")

    # Получаем только города России
    russia_cities = []
    for city_name, city_info in hh_areas.items():
        if city_info.get('root_parent_id') == '113':
            russia_cities.append(city_name)

    # Мультиселект для выбора городов
    selected_cities = st.multiselect(
        "Выберите город(а) для проверки и выгрузки:",
        options=sorted(russia_cities),
        key="geo_checker",
        help="Выберите один или несколько городов"
    )

    # Показываем информацию о выбранных городах
    if selected_cities:
        st.markdown(f"**Выбрано городов:** {len(selected_cities)}")

        # Создаем DataFrame для выбранных городов
        selected_cities_data = []
        for city_name in selected_cities:
            city_info = hh_areas[city_name]
            selected_cities_data.append({
                'Город': city_name,
                'ID HH': city_info['id'],
                'Регион': city_info['parent']
            })

        selected_cities_df = pd.DataFrame(selected_cities_data)
        st.dataframe(selected_cities_df, use_container_width=True, hide_index=True)

        # Кнопка выгрузки выбранных городов
        col1, col2 = st.columns(2)
        with col1:
            # Для публикатора (только названия городов)
            publisher_df = pd.DataFrame({'Город': selected_cities_df['Город']})
            # Санитизация данных перед экспортом (защита от CSV Injection)
            publisher_df = sanitize_csv_content(publisher_df)
            output_pub = io.BytesIO()
            with pd.ExcelWriter(output_pub, engine='openpyxl') as writer:
                publisher_df.to_excel(writer, index=False, header=False, sheet_name='Гео')
            output_pub.seek(0)
            st.download_button(
                label=f"📤 Для публикатора ({len(selected_cities)} городов)",
                data=output_pub,
                file_name="selected_cities_publisher.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                key="download_selected_publisher"
            )
        with col2:
            # Полный отчет с ID и регионами
            # Санитизация данных перед экспортом (защита от CSV Injection)
            safe_cities_df = sanitize_csv_content(selected_cities_df.copy())
            output_full = io.BytesIO()
            with pd.ExcelWriter(output_full, engine='openpyxl') as writer:
                safe_cities_df.to_excel(writer, index=False, sheet_name='Города')
            output_full.seek(0)
            st.download_button(
                label=f"📥 Полный отчет ({len(selected_cities)} городов)",
                data=output_full,
                file_name="selected_cities.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                key="download_selected_full"
            )

    # КНОПКА ВЫГРУЗКИ ВСЕХ ГОРОДОВ
    st.markdown("")
    if st.button("🌍 Выгрузить ВСЕ города из справочника", type="secondary", use_container_width=False, key="export_all_cities_btn"):
        with st.spinner("Формирую полный список..."):
            all_cities_df = get_all_cities(hh_areas)
            if not all_cities_df.empty:
                st.success(f"✅ Найдено **{len(all_cities_df)}** городов в справочнике HH.ru")
                st.dataframe(all_cities_df, use_container_width=True, height=400)

                col1, col2 = st.columns(2)
                with col1:
                    publisher_df = pd.DataFrame({'Город': all_cities_df['Город']})
                    # Санитизация данных перед экспортом (защита от CSV Injection)
                    publisher_df = sanitize_csv_content(publisher_df)
                    output_pub = io.BytesIO()
                    with pd.ExcelWriter(output_pub, engine='openpyxl') as writer:
                        publisher_df.to_excel(writer, index=False, header=False, sheet_name='Гео')
                    output_pub.seek(0)
                    st.download_button(
                        label=f"📤 Для публикатора ({len(all_cities_df)} городов)",
                        data=output_pub,
                        file_name="all_cities_publisher.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key="download_all_publisher"
                    )
                with col2:
                    # Санитизация данных перед экспортом (защита от CSV Injection)
                    safe_all_cities_df = sanitize_csv_content(all_cities_df.copy())
                    output_full = io.BytesIO()
                    with pd.ExcelWriter(output_full, engine='openpyxl') as writer:
                        safe_all_cities_df.to_excel(writer, index=False, sheet_name='Города')
                    output_full.seek(0)
                    st.download_button(
                        label=f"📥 Скачать полный отчет ({len(all_cities_df)} городов)",
                        data=output_full,
                        file_name="all_cities.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key="download_all_full"
                    )

st.markdown("---")

# ============================================
# БЛОК: СИНХРОНИЗАТОР ГОРОДОВ
# ============================================
st.markdown('<div id="синхронизатор-городов"></div>', unsafe_allow_html=True)
st.markdown("## **📤 Синхронизатор городов**")

with st.sidebar:
    # OPTIMIZED: use cached logo loading
    try:
        logo_base64 = get_cached_icon_base64("min-hh-red.png")

        if logo_base64:
            # Вставляем через HTML с прямыми стилями для максимального качества
            st.markdown(
                f'''<img src="{logo_base64}"
                style="width: 200px;
                       height: auto;
                       image-rendering: auto;
                       -ms-interpolation-mode: bicubic;
                       display: block;
                       margin-bottom: 10px;
                       object-fit: contain;" />''',
                unsafe_allow_html=True
            )
        else:
            logger.warning("Не удалось загрузить min-hh-red.png")
            st.markdown(f'<div class="title-container"><span>{SYNC_ICON}</span></div>', unsafe_allow_html=True)
    except Exception as e:
        # Fallback если PNG еще не создан
        st.markdown(
            f'<div class="title-container">'
            f'<span class="rotating-earth">{SYNC_ICON}</span>'
            f'</div>',
            unsafe_allow_html=True
        )
    st.markdown("---")

    # Инициализация состояния для отображения инструкций
    if 'show_instruction' not in st.session_state:
        st.session_state.show_instruction = None

    # Словарь с инструкциями для каждого раздела
    instructions = {
        "проверка-гео": """
### Проверка гео и выгрузка базы

<p><span class="step-number">1</span> <strong>Быстрая проверка города</strong></p>

- Введите название города в поисковое поле
- Система покажет ID города в HH.ru и его регион
- Используйте для быстрой проверки наличия города в справочнике

<p><span class="step-number">2</span> <strong>Выгрузка всех городов</strong></p>

- Нажмите кнопку "Выгрузить ВСЕ города"
- Получите Excel-файл со всеми городами России из справочника HH.ru
- Файл содержит: название города, ID, регион, тип населенного пункта
        """,

        "синхронизатор-городов": """
### Синхронизатор городов

<p><span class="step-number">1</span> <strong>Простой сценарий (один столбец)</strong></p>

- Загрузите файл, где в первом столбце указаны города
- Система автоматически сопоставит города со справочником HH.ru
- Подходит для быстрой проверки списка городов

<p><span class="step-number">2</span> <strong>Сценарий со столбцом "Вакансия"</strong></p>

- Загрузите файл с колонкой "Вакансия"
- Данные будут разделены по вакансиям и обработаны отдельно
- Скачайте результат единым архивом или отдельными файлами

<p><span class="step-number">3</span> <strong>Сценарий с вкладками "вакансия"</strong></p>

- Загрузите Excel с несколькими вкладками (названия начинаются на "вакансия")
- Каждая вкладка обрабатывается как отдельная вакансия
- Идеально для структурированной работы с множеством вакансий

**Порядок работы:**
1. Загрузите файл → 2. Выберите режим → 3. Нажмите "Начать сопоставление" → 4. Проверьте и отредактируйте → 5. Скачайте
        """,

        "выбор-регионов-и-городов": """
### Выбор регионов и городов

<p><span class="step-number">1</span> <strong>Поиск по регионам</strong></p>

- Выберите регионы из списка (можно несколько)
- Получите Excel-файл со всеми городами выбранных регионов
- Файл содержит полную информацию о городах

<p><span class="step-number">2</span> <strong>Поиск по городам</strong></p>

- Выберите конкретные города (можно несколько)
- Получите информацию: ID, название, регион
- Скачайте в формате Excel

<p><span class="step-number">3</span> <strong>Поиск по населению</strong></p>

- Укажите минимальное и максимальное население
- Система найдет все города в указанном диапазоне
- Данные о населении из актуального справочника
        """,

        "объединитель-файлов": """
### Объединитель файлов

<p><span class="step-number">1</span> <strong>Загрузите файлы</strong></p>

- Поддерживаются форматы: Excel (xlsx, xls, xlsm, xlsb) и CSV
- Можно загрузить несколько файлов одновременно
- Все файлы должны иметь одинаковую структуру столбцов

<p><span class="step-number">2</span> <strong>Обработка</strong></p>

- Система автоматически объединит все файлы
- Полные дубликаты будут выделены оранжевым цветом
- Дубликаты размещаются в начале файла для удобства

<p><span class="step-number">3</span> <strong>Скачивание</strong></p>

- Нажмите кнопку "Скачать объединенный файл"
- Файл содержит статистику: общее количество, дубликаты, уникальные записи
        """,

        "сверки-с-клиентами": """
### Сверки с клиентами

<p><span class="step-number">1</span> <strong>Сверка Я.Еда</strong></p>

- Нажмите на желтую карточку "Яндекс.Еда"
- Скопируйте код установки библиотек (Блок 1)
- Запустите его в Google Colab и дождитесь завершения

<p><span class="step-number">2</span> <strong>Основной код</strong></p>

- Скопируйте основной код сверки (Блок 2)
- Вставьте в новую ячейку Google Colab
- Запустите и следуйте инструкциям на экране

<p><span class="step-number">3</span> <strong>Файлы</strong></p>

Подготовьте файлы с названиями:
- "ООО Хэдхантер Биллинг....." (отчет биллинг)
- "Отчет-по-откликам-по-проектам-работодателя-" (внутренний отчет HH)
- "Leads_" (лиды из ЛК Я.Еды)

⏱️ Время выполнения: 30-40 минут
        """
    }

    st.markdown("### 🧭 Навигация")

    # Якорная навигация (стили в static/styles.css)
    nav_items = [
        ("Проверка гео и выгрузка базы", "проверка-гео"),
        ("Синхронизатор городов", "синхронизатор-городов"),
        ("Выбор регионов и городов", "выбор-регионов-и-городов"),
        ("Объединитель файлов", "объединитель-файлов"),
        ("Сверки с клиентами", "сверки-с-клиентами")
    ]

    for name, anchor in nav_items:
        st.markdown(f'<a class="nav-link" href="#{anchor}">{name}</a>', unsafe_allow_html=True)

    st.markdown("---")

    # Инструкции в раскрывающихся блоках
    st.markdown("### 📖 Инструкции")

    with st.expander("Проверка гео и выгрузка базы"):
        st.markdown(instructions["проверка-гео"], unsafe_allow_html=True)

    with st.expander("Синхронизатор городов"):
        st.markdown(instructions["синхронизатор-городов"], unsafe_allow_html=True)

    with st.expander("Выбор регионов и городов"):
        st.markdown(instructions["выбор-регионов-и-городов"], unsafe_allow_html=True)

    with st.expander("Объединитель файлов"):
        st.markdown(instructions["объединитель-файлов"], unsafe_allow_html=True)

    with st.expander("Сверки с клиентами"):
        st.markdown(instructions["сверки-с-клиентами"], unsafe_allow_html=True)

    st.markdown("---")

# Устанавливаем порог совпадения как константу
threshold = 85

# ============================================
# ЗАГРУЗКА И ОБРАБОТКА ФАЙЛОВ
# ============================================
st.subheader("📁 Загрузка файлов")
uploaded_files = st.file_uploader(
    "Выберите один или несколько файлов с городами",
    type=['xlsx', 'csv'],
    help="Поддерживаются форматы: Excel (.xlsx) и CSV. Можно загрузить несколько файлов одновременно",
    accept_multiple_files=True,
    key="files_uploader"
)

if uploaded_files and hh_areas is not None:
    st.markdown("---")

    # Валидация размера и расширения файлов
    files_valid = True
    for uploaded_file in uploaded_files:
        # Проверка размера
        is_valid_size, error_msg = validate_file_size(uploaded_file.size)
        if not is_valid_size:
            st.error(f"❌ {uploaded_file.name}: {error_msg}")
            logger.warning(f"Файл отклонен (размер): {uploaded_file.name} ({uploaded_file.size} байт)")
            log_security_event('file_size_exceeded', f"{uploaded_file.name}: {uploaded_file.size} байт", 'WARNING')
            files_valid = False

        # Проверка расширения
        is_valid_ext, error_msg = validate_file_extension(uploaded_file.name, ['.xlsx', '.csv'])
        if not is_valid_ext:
            st.error(f"❌ {uploaded_file.name}: {error_msg}")
            logger.warning(f"Файл отклонен (расширение): {uploaded_file.name}")
            log_security_event('invalid_file_extension', uploaded_file.name, 'WARNING')
            files_valid = False

    if not files_valid:
        st.stop()

    try:
        # Обрабатываем все загруженные файлы
        sheets_data = {}
        file_counter = 1

        for uploaded_file in uploaded_files:
            # Определяем тип файла и читаем все вкладки
            if uploaded_file.name.endswith('.csv'):
                # CSV - одна вкладка
                df = pd.read_csv(uploaded_file, header=None)
                # Если несколько файлов, добавляем префикс к имени
                sheet_key = f"Файл{file_counter}_Sheet1" if len(uploaded_files) > 1 else "Sheet1"
                sheets_data[sheet_key] = df
            else:
                # Excel - читаем все вкладки
                excel_file = pd.ExcelFile(uploaded_file)
                for sheet_name in excel_file.sheet_names:
                    df_sheet = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    if len(df_sheet) > 0:  # Только непустые вкладки
                        # Если несколько файлов, добавляем префикс к имени вкладки
                        sheet_key = f"Файл{file_counter}_{sheet_name}" if len(uploaded_files) > 1 else sheet_name
                        sheets_data[sheet_key] = df_sheet
            file_counter += 1
        
        # Анализируем структуру файла
        st.session_state.sheets_data = {}
        st.session_state.has_multiple_sheets = len(sheets_data) > 1
        st.session_state.sheet_mode = None  # 'tabs' или 'columns' или 'both' или None
        
        # Обрабатываем каждую вкладку
        for sheet_name, df in sheets_data.items():
            has_header = False
            has_vacancy_column = False
            vacancy_col_idx = None
            
            # Проверяем первую строку на наличие заголовков
            if len(df) > 0:
                first_row = df.iloc[0]
                # Проверяем первую ячейку на "Город"
                if pd.notna(first_row[0]) and 'город' in str(first_row[0]).lower():
                    has_header = True
                    # Ищем столбец "Вакансия"
                    for idx, val in enumerate(first_row):
                        if pd.notna(val) and 'вакансия' in str(val).lower():
                            has_vacancy_column = True
                            vacancy_col_idx = idx
                            break
            
            # Если есть заголовок, делаем его названиями столбцов
            if has_header:
                df.columns = df.iloc[0]
                df = df.iloc[1:].reset_index(drop=True)
            
            # Сохраняем данные вкладки
            st.session_state.sheets_data[sheet_name] = {
                'df': df.copy(),
                'has_vacancy_column': has_vacancy_column,
                'vacancy_col_idx': vacancy_col_idx
            }
        
        # Определяем режим работы
        if st.session_state.has_multiple_sheets:
            # Проверяем есть ли вкладки с "вакансия" в названии
            vacancy_sheets = [name for name in sheets_data.keys() 
                            if 'вакансия' in name.lower() or 'вакансии' in name.lower()]
            
            # Проверяем есть ли столбцы "Вакансия" в каких-то вкладках
            sheets_with_vacancy_column = [name for name, data in st.session_state.sheets_data.items() 
                                         if data['has_vacancy_column']]
            
            if vacancy_sheets or len(st.session_state.sheets_data) > 1:
                # Есть вкладки - режим вкладок
                st.session_state.sheet_mode = 'tabs'
                
                # Если еще и столбцы есть - комбинированный режим
                if sheets_with_vacancy_column:
                    st.session_state.sheet_mode = 'both'
                    
                files_info = f" из **{len(uploaded_files)}** файлов" if len(uploaded_files) > 1 else ""
                st.info(f"📄 Загружено **{len(sheets_data)}** вкладок{files_info} | 🎯 **Обнаружен режим работы с вкладками**")
            else:
                st.session_state.sheet_mode = None
                files_info = f" из **{len(uploaded_files)}** файлов" if len(uploaded_files) > 1 else ""
                st.info(f"📄 Загружено **{len(sheets_data)}** вкладок{files_info}")
        else:
            # Одна вкладка - проверяем столбец "Вакансия"
            first_sheet_data = list(st.session_state.sheets_data.values())[0]
            if first_sheet_data['has_vacancy_column']:
                st.session_state.sheet_mode = 'columns'
        
        # Для обратной совместимости - сохраняем первую вкладку как основной DF
        first_sheet_name = list(sheets_data.keys())[0]
        st.session_state.original_df = st.session_state.sheets_data[first_sheet_name]['df'].copy()
        st.session_state.has_vacancy_mode = st.session_state.sheet_mode in ['columns', 'tabs', 'both']

        # Показываем превью файла с информацией о размерах
        vacancy_info = " | 🎯 **Обнаружен столбец 'Вакансия'**" if has_vacancy_column else ""
        with st.expander(f"👀 Превью ({len(df)} строк, {len(df.columns)} столбцов{vacancy_info})", expanded=False):
            if st.session_state.has_multiple_sheets:
                # Показываем вкладки для выбора
                sheet_tabs = st.tabs(list(st.session_state.sheets_data.keys()))
                for tab, sheet_name in zip(sheet_tabs, st.session_state.sheets_data.keys()):
                    with tab:
                        st.dataframe(st.session_state.sheets_data[sheet_name]['df'].head(), use_container_width=True)
            else:
                # Одна вкладка
                st.dataframe(st.session_state.original_df.head(), use_container_width=True)
          
        if st.button("🚀 Начать сопоставление", type="primary", use_container_width=True):  
            with st.spinner("Обрабатываю..."):  
                # Обрабатываем каждую вкладку
                st.session_state.sheets_results = {}
                
                for sheet_name, sheet_data in st.session_state.sheets_data.items():
                    df_sheet = sheet_data['df']
                    result_df, dup_original, dup_hh, total_dup = match_cities(df_sheet, hh_areas, threshold, sheet_name=sheet_name)  
                    
                    st.session_state.sheets_results[sheet_name] = {
                        'result_df': result_df,
                        'dup_original': dup_original,
                        'dup_hh': dup_hh,
                        'total_dup': total_dup,
                        'has_vacancy_column': sheet_data['has_vacancy_column']
                    }
                
                # Для обратной совместимости - сохраняем первую вкладку
                first_sheet = list(st.session_state.sheets_results.keys())[0]
                st.session_state.result_df = st.session_state.sheets_results[first_sheet]['result_df']
                st.session_state.dup_original = st.session_state.sheets_results[first_sheet]['dup_original']
                st.session_state.dup_hh = st.session_state.sheets_results[first_sheet]['dup_hh']
                st.session_state.total_dup = st.session_state.sheets_results[first_sheet]['total_dup']
                
                st.session_state.processed = True
                st.session_state.manual_selections = {}
                st.session_state.search_query = ""
                st.session_state.added_cities = []
                st.session_state.candidates_cache = {}

                # Очищаем кэши функций
                apply_manual_selections_cached.clear()

                # Очищаем кэши от предыдущих файлов
                if 'vacancy_files' in st.session_state:
                    del st.session_state.vacancy_files
          
        if st.session_state.processed and st.session_state.result_df is not None:
            # Прямая ссылка вместо .copy() - копируем только при изменении
            result_df = st.session_state.result_df  
            dup_original = st.session_state.dup_original
            dup_hh = st.session_state.dup_hh
            total_dup = st.session_state.total_dup

            # Стандартные блоки показываем только если НЕТ режима вакансий
            # Если есть вакансии - используем специальные блоки для split или single
            show_standard_blocks = not st.session_state.get('has_vacancy_mode', False)
            
            if show_standard_blocks:
                st.markdown("---")  
                st.subheader("📊 Результаты")  
                  
                col1, col2, col3, col4, col5, col6 = st.columns(6)  
                  
                total = len(result_df)  
                exact = len(result_df[result_df['Статус'] == '✅ Точное'])  
                similar = len(result_df[result_df['Статус'] == '⚠️ Похожее'])  
                duplicates = len(result_df[result_df['Статус'].str.contains('Дубликат', na=False)])  
                not_found = len(result_df[result_df['Статус'] == '❌ Не найдено'])  
                  
                to_export = len(result_df[  
                    (~result_df['Статус'].str.contains('Дубликат', na=False)) &   
                    (result_df['Итоговое гео'].notna())  
                ])  
                  
                col1.metric("Всего", total)  
                col2.metric("✅ Точных", exact)  
                col3.metric("⚠️ Похожих", similar)  
                col4.metric("🔄 Дубликатов", duplicates)  
                col5.metric("❌ Не найдено", not_found)  
                col6.metric("📤 К выгрузке", to_export)  
                  
                if duplicates > 0:
                    st.warning(f"""
                    ⚠️ **Найдено {duplicates} дубликатов:**
                    - 🔄 По исходному названию: **{dup_original}**
                    - 🔄 По результату HH: **{dup_hh}**
                    """)

                # Проверяем наличие гео из других стран (VECTORIZED - быстрее в ~100 раз!)
                russia_id = '113'
                non_russian_cities = []

                # Фильтруем только строки с валидным гео
                valid_geo_mask = result_df['Итоговое гео'].notna()
                if valid_geo_mask.any():
                    valid_rows = result_df[valid_geo_mask]

                    # Векторизованная проверка принадлежности к России
                    for geo_name in valid_rows['Итоговое гео'].unique():
                        if geo_name in hh_areas:
                            city_info = hh_areas[geo_name]
                            if city_info.get('root_parent_id', '') != russia_id:
                                # Находим все строки с этим городом
                                city_rows = valid_rows[valid_rows['Итоговое гео'] == geo_name]
                                for original in city_rows['Исходное название'].unique():
                                    non_russian_cities.append({
                                        'original': original,
                                        'matched': geo_name,
                                        'country_id': city_info.get('root_parent_id', 'Unknown')
                                    })

                if non_russian_cities:
                    st.error(f"""
                    🌍 **Обнаружено {len(non_russian_cities)} гео из других стран!**

                    Эти города не из России и не должны попадать в выгрузку.
                    Пожалуйста, проверьте и исправьте совпадения ниже в блоке редактирования.
                    """)

                    # Показываем список
                    with st.expander("🔍 Показать гео из других стран"):
                        for city in non_russian_cities:
                            st.text(f"• {city['original']} → {city['matched']}")

            # РАННЯЯ ОСТАНОВКА ДЛЯ РЕЖИМА SPLIT И SINGLE С ВАКАНСИЯМИ
            # Если режим split или single с вакансиями - пропускаем все стандартные блоки и сразу переходим к специальным блокам
            if st.session_state.get('has_vacancy_mode', False):
                # Переход к блоку "ПРОВЕРЯЕМ РЕЖИМ РАБОТЫ" ниже
                # Там будут показаны специальные блоки редактирования для split или single режимов
                pass
            else:
                # Для обычного режима БЕЗ вакансий показываем стандартные блоки

                    st.markdown("---")
                    st.subheader("📋 Таблица сопоставлений")

                    # Поле поиска и фильтры в двух колонках
                    col_search, col_status = st.columns([2, 1])

                    with col_search:
                        st.text_input(
                            "🔍 Поиск по таблице",
                            key="search_query",
                            placeholder="Начните вводить название города...",
                            label_visibility="visible"
                        )

                    with col_status:
                        # Определяем доступные статусы
                        available_statuses = result_df['Статус'].unique().tolist()
                        status_filter = st.multiselect(
                            "📊 Фильтр по статусам",
                            options=available_statuses,
                            default=[],
                            key="status_filter",
                            label_visibility="visible"
                        )

                    # VECTORIZED: sort priority (0=no match, 1=changed, 2=unchanged)
                    result_df['sort_priority'] = np.where(
                        result_df['Совпадение %'] == 0, 0,
                        np.where(result_df['Изменение'] == 'Да', 1, 2)
                    )

                    result_df_sorted = result_df.sort_values(
                        by=['sort_priority', 'Совпадение %'],
                        ascending=[True, True]
                    ).reset_index(drop=True)

                    # Применяем фильтр по статусам
                    if status_filter:
                        result_df_sorted = result_df_sorted[result_df_sorted['Статус'].isin(status_filter)]

                    if st.session_state.search_query and st.session_state.search_query.strip():
                        # Sanitization пользовательского ввода для защиты от инъекций
                        sanitized_query = sanitize_user_input(st.session_state.search_query, max_length=200)
                        search_lower = sanitized_query.lower().strip()
                        # VECTORIZED: search mask across multiple columns
                        mask = (
                            result_df_sorted['Исходное название'].astype(str).str.lower().str.contains(search_lower, na=False) |
                            result_df_sorted['Итоговое гео'].astype(str).str.lower().str.contains(search_lower, na=False) |
                            result_df_sorted['Регион'].astype(str).str.lower().str.contains(search_lower, na=False) |
                            result_df_sorted['Статус'].astype(str).str.lower().str.contains(search_lower, na=False)
                        )
                        result_df_filtered = result_df_sorted[mask]

                        if len(result_df_filtered) == 0:
                            st.warning(f"По запросу **'{sanitized_query}'** ничего не найдено")
                        else:
                            st.info(f"Найдено совпадений: **{len(result_df_filtered)}** из {len(result_df)}")
                    else:
                        result_df_filtered = result_df_sorted  
              
                    display_df = result_df_filtered.copy()
                    display_df = display_df.drop(['row_id', 'sort_priority'], axis=1, errors='ignore')

                    # Сбрасываем индекс чтобы избежать дублирования
                    display_df = display_df.reset_index(drop=True)

                    st.dataframe(display_df, use_container_width=True, height=400, hide_index=True)  
              
                    # ИЗМЕНЕНО: Исключаем дубликаты из редактирования, порог 95%
                    editable_rows = result_df_sorted[
                        (result_df_sorted['Совпадение %'] <= 95) &
                        (~result_df_sorted['Статус'].str.contains('Дубликат', na=False))
                    ].copy()

                    # Сортируем: сначала "Нет совпадения", затем по возрастанию процента
                    if len(editable_rows) > 0:
                        # Создаем приоритет: 0 для "Нет совпадения", 1 для остальных
                        # VECTORIZED: sort priority (0 for not found, 1 for others)
                        editable_rows['_sort_priority'] = (~editable_rows['Статус'].str.contains('❌ Не найдено', na=False)).astype(int)
                        editable_rows = editable_rows.sort_values(
                            ['_sort_priority', 'Совпадение %'],
                            ascending=[True, True]
                        )
                        editable_rows = editable_rows.drop(columns=['_sort_priority'])  
              
                    if len(editable_rows) > 0:
                        st.markdown("---")
                        st.subheader("✏️ Редактирование городов с совпадением ≤ 95%")

                        # Callback для сохранения выбора ТОЛЬКО при изменении
                        def on_city_select_scenario1(row_id, widget_key):
                            """Callback для сценария 1 - вызывается только при изменении"""
                            selected = st.session_state.get(widget_key)
                            if selected == "❌ Нет совпадения":
                                st.session_state.manual_selections[row_id] = "❌ Нет совпадения"
                            elif selected:
                                # Извлекаем название без процента
                                city_match = selected.rsplit(' (', 1)[0]
                                st.session_state.manual_selections[row_id] = city_match

                        for idx, row in editable_rows.iterrows():
                            with st.container():
                                row_id = row['row_id']
                                city_name = row['Исходное название']
                                current_value = row['Итоговое гео']
                                current_match = row['Совпадение %']

                                # Получаем кандидатов из кэша или вычисляем
                                candidates = st.session_state.candidates_cache.get(row_id, [])
                                if not candidates:
                                    candidates = get_candidates_by_word(city_name, get_russian_cities_cached(hh_areas), limit=20)

                                # Кэшированная подготовка options (избегаем повторных вычислений)
                                options, candidates_dict = prepare_city_options(
                                    tuple(candidates),  # tuple для кэширования
                                    current_value,
                                    current_match,
                                    city_name
                                )

                                # Определяем выбранное значение
                                widget_key = f"select_{row_id}"
                                if row_id in st.session_state.manual_selections:
                                    selected_value = st.session_state.manual_selections[row_id]
                                else:
                                    selected_value = current_value

                                # Быстрый поиск индекса O(1) вместо O(n)
                                if selected_value == "❌ Нет совпадения":
                                    default_idx = 0
                                else:
                                    default_idx = candidates_dict.get(selected_value, 0)

                                col1, col2, col3, col4 = st.columns([2, 3, 1, 1])

                                with col1:
                                    st.markdown(f"**{row['Исходное название']}**")

                                with col2:
                                    st.selectbox(
                                        "Выберите город:",
                                        options=options,
                                        index=default_idx,
                                        key=widget_key,
                                        label_visibility="collapsed",
                                        on_change=on_city_select_scenario1,
                                        args=(row_id, widget_key)
                                    )

                                with col3:
                                    st.text(f"{row['Совпадение %']}%")

                                with col4:
                                    st.text(row['Статус'])

                                st.markdown("<hr style='margin-top: 5px; margin-bottom: 5px;'>", unsafe_allow_html=True)

                        # ============================================
                        # БЛОК: ДОБАВЛЕНИЕ ЛЮБОГО ГОРОДА (только для НЕ split режима)
                        # ============================================
                        st.markdown("---")
                        st.subheader("➕ Добавить дополнительные города")
                
                        # Селектор на половину ширины экрана
                        col_selector = st.columns([1, 1])
                        with col_selector[0]:
                            # Используем кэшированную версию вместо цикла
                            russia_cities = get_russian_cities_cached(hh_areas)

                            selected_city = st.selectbox(
                                "Выберите город:",
                                options=sorted(russia_cities),
                                key="city_selector",
                                help="Выберите город из справочника HH.ru"
                            )

                        # Кнопки под селектором
                        col_btn1, col_btn2 = st.columns(2)
                        with col_btn1:
                            if st.button("➕ Добавить", use_container_width=True, type="primary"):
                                if selected_city and selected_city not in st.session_state.added_cities:
                                    st.session_state.added_cities.append(selected_city)
                                    st.success(f"✅ {selected_city}")
                                elif selected_city in st.session_state.added_cities:
                                    st.warning(f"⚠️ Уже добавлен")

                        with col_btn2:
                            if st.button("🗑️ Очистить", use_container_width=True):
                                st.session_state.added_cities = []
                                st.rerun()
                
                        # Показываем список добавленных городов
                        if st.session_state.added_cities:
                            st.success(f"📋 Добавлено городов: **{len(st.session_state.added_cities)}**")
                    
                            # Показываем города в компактном виде
                            added_cities_text = ", ".join(st.session_state.added_cities)
                            st.text_area(
                                "Список добавленных городов:",
                                value=added_cities_text,
                                height=100,
                                disabled=True,
                                label_visibility="collapsed"
                            )

                    # Если режим split - переходим сразу к блоку редактирования по вакансиям, пропуская стандартные блоки скачивания
                    if not show_standard_blocks:
                        # Режим split или single с вакансиями - пропускаем весь блок скачивания и идем к вакансиям
                        pass
                    else:
                        # Обычный режим БЕЗ вакансий - показываем блок скачивания
                        st.markdown("---")
                        st.subheader("💾 Скачать результаты")

                        # КЭШИРОВАННОЕ применение ручных изменений
                        # Выполняется ТОЛЬКО при изменении manual_selections, а не при каждом rerun!
                        # Было: ~1000ms при каждом клике → Стало: ~5ms (берется из кэша)
                        final_result_df = apply_manual_selections_cached(
                            result_df,
                            st.session_state.manual_selections,
                            hh_areas,
                            cache_key="scenario1"
                        )

                        # Добавляем города из added_cities
                        if st.session_state.added_cities:
                            for city in st.session_state.added_cities:
                                if city in hh_areas:
                                    final_result_df = pd.concat([final_result_df, pd.DataFrame([{
                                        'row_id': len(final_result_df),
                                        'Исходное название': city,
                                        'Итоговое гео': city,
                                        'ID HH': hh_areas[city]['id'],
                                        'Регион': hh_areas[city]['parent'],
                                        'Совпадение %': 100.0,
                                        'Статус': '✅ Добавлено',
                                        'Изменение': 'Нет'
                                    }])], ignore_index=True)  
            
            # ПРОВЕРЯЕМ РЕЖИМ РАБОТЫ
            # Если есть вакансии - показываем блок редактирования по вакансиям/вкладкам
            if st.session_state.get('has_vacancy_mode', False):
                # РЕЖИМ: Разделение по вакансиям/вкладкам с редактированием
                st.markdown("---")
                
                # Определяем тип разделения: по вкладкам или по столбцу вакансий
                if st.session_state.sheet_mode == 'tabs':
                    # РЕЖИМ ВКЛАДОК: каждая вкладка = отдельный файл
                    st.subheader("🎯 Редактирование и выгрузка по вкладкам")
                    
                    # Получаем список вкладок
                    sheet_names = list(st.session_state.sheets_results.keys())
                    st.success(f"📊 Найдено **{len(sheet_names)}** вкладок")
                    
                    # Инициализируем состояние
                    if 'vacancy_files' not in st.session_state:
                        st.session_state.vacancy_files = {}

                    # Инициализируем выбранную вкладку
                    if 'selected_sheet' not in st.session_state:
                        st.session_state.selected_sheet = sheet_names[0]

                    # Выбор вкладки через radio кнопки (как в Сценарии 2)
                    st.markdown("#### 📋 Выберите вкладку для редактирования:")
                    selected_sheet = st.radio(
                        "Вкладка:",
                        options=sheet_names,
                        index=sheet_names.index(st.session_state.selected_sheet) if st.session_state.selected_sheet in sheet_names else 0,
                        key="sheet_selector",
                        horizontal=False,
                        label_visibility="collapsed"
                    )
                    st.session_state.selected_sheet = selected_sheet

                    st.markdown("---")

                    # Показываем контент только для выбранной вкладки
                    sheet_name = selected_sheet
                    tab_idx = sheet_names.index(sheet_name)

                    st.markdown(f"### 📄 {sheet_name}")
                    
                    # Получаем данные этой вкладки
                    sheet_result = st.session_state.sheets_results[sheet_name]
                    result_df_sheet = sheet_result['result_df']
                    original_df_sheet = st.session_state.sheets_data[sheet_name]['df']

                    # Блок редактирования городов с совпадением ≤ 95%
                    editable_rows = result_df_sheet[
                        (result_df_sheet['Совпадение %'] <= 95) &
                        (~result_df_sheet['Статус'].str.contains('Дубликат', na=False))
                    ].copy()
                    
                    if len(editable_rows) > 0:
                        # Убираем дубликаты по исходному названию (VECTORIZED)
                        editable_rows['_normalized_original'] = (
                            editable_rows['Исходное название']
                            .fillna('').astype(str)
                            .str.replace('ё', 'е').str.replace('Ё', 'Е')
                            .str.lower().str.strip()
                            .str.replace(r'\s+', ' ', regex=True)
                        )
                        editable_rows = editable_rows.drop_duplicates(subset=['_normalized_original'], keep='first')

                        # Сортируем: сначала "Нет совпадения", затем по возрастанию процента
                        # VECTORIZED: sort priority (0 for not found, 1 for others)
                        editable_rows['_sort_priority'] = (~editable_rows['Статус'].str.contains('❌ Не найдено', na=False)).astype(int)
                        editable_rows = editable_rows.sort_values(
                            ['_sort_priority', 'Совпадение %'],
                            ascending=[True, True]
                        )
                        editable_rows = editable_rows.drop(columns=['_sort_priority'])

                        st.markdown("#### ✏️ Редактирование городов с совпадением ≤ 95%")

                        # ============================================
                        # CALLBACK для предотвращения полного rerun
                        # ============================================
                        def on_city_select_tab(selection_key, widget_key):
                            """Callback для режима split - вызывается только при изменении"""
                            selected = st.session_state.get(widget_key)
                            if selected == "❌ Нет совпадения":
                                st.session_state.manual_selections[selection_key] = "❌ Нет совпадения"
                            elif selected:
                                # Извлекаем название без процента
                                city_match = selected.rsplit(' (', 1)[0]
                                st.session_state.manual_selections[selection_key] = city_match

                        # ============================================
                        # Для каждого города показываем выбор
                        for idx, row in editable_rows.iterrows():
                            row_id = row['row_id']
                            city_name = row['Исходное название']
                            current_value = row['Итоговое гео']
                            current_match = row['Совпадение %']

                            # Используем кэш кандидатов из smart_match_city
                            cache_key = (sheet_name, row_id)
                            candidates = st.session_state.candidates_cache.get(cache_key, [])
                            if not candidates:
                                candidates = get_candidates_by_word(city_name, get_russian_cities_cached(hh_areas), limit=20)

                            # Кэшированная подготовка options (избегаем повторных вычислений)
                            options, candidates_dict = prepare_city_options(
                                tuple(candidates),
                                current_value,
                                current_match,
                                city_name
                            )

                            # Определяем текущий выбор
                            unique_key = f"select_{sheet_name}_{row_id}_{tab_idx}"
                            selection_key = (sheet_name, row_id)

                            if selection_key in st.session_state.manual_selections:
                                selected_value = st.session_state.manual_selections[selection_key]
                            else:
                                selected_value = current_value

                            # Быстрый поиск индекса O(1)
                            if selected_value == "❌ Нет совпадения":
                                default_idx = 0
                            else:
                                default_idx = candidates_dict.get(selected_value, 0)

                            col1, col2, col3 = st.columns([2, 3, 1])

                            with col1:
                                st.text(city_name)

                            with col2:
                                st.selectbox(
                                    "Выберите город:",
                                    options=options,
                                    index=default_idx,
                                    key=unique_key,
                                    label_visibility="collapsed",
                                    on_change=on_city_select_tab,
                                    args=(selection_key, unique_key)
                                )

                            with col3:
                                st.text(f"{row['Совпадение %']:.1f}%")

                            # VISUAL: Добавляем разделитель как в Сценарии 2
                            st.markdown("<hr style='margin-top: 5px; margin-bottom: 5px;'>", unsafe_allow_html=True)

                    # Применяем ручные изменения через КЭШИРОВАННУЮ функцию
                    # Фильтруем только изменения для текущей вкладки
                    sheet_selections = {}
                    for selection_key, new_value in st.session_state.manual_selections.items():
                        if isinstance(selection_key, tuple):
                            key_sheet_name, row_id = selection_key
                            if key_sheet_name == sheet_name:
                                sheet_selections[row_id] = new_value
                        else:
                            # Для обратной совместимости
                            sheet_selections[selection_key] = new_value

                    # Используем кэшированную функцию вместо цикла
                    # FIX: Передаем sheet_name в cache_key для уникальности кэша каждой вкладки
                    result_df_sheet_final = apply_manual_selections_cached(
                        result_df_sheet,
                        sheet_selections,
                        hh_areas,
                        cache_key=f"tab_{sheet_name}"
                    )
                    
                    # FIX: Используем кэшированную функцию подготовки данных
                    # Это предотвращает повторные вычисления (merge, filter) при каждом клике
                    final_output = prepare_final_sheet_output_cached(
                        result_df_sheet_final,
                        original_df_sheet,
                        sheet_name
                    )

                    if len(final_output) > 0:

                        # Превью итогового файла для вкладки
                        st.markdown(f"#### 👀 Превью итогового файла - {sheet_name}")
                        st.dataframe(final_output, use_container_width=True, height=300)

                        # Кнопка скачивания
                        st.markdown("---")
                        safe_sheet_name = str(sheet_name).replace('/', '_').replace('\\', '_')[:50]

                        # OPTIMIZED: Используем кэшированную генерацию файла
                        excel_bytes = create_excel_bytes_cached(final_output, sheet_name)
                        
                        st.download_button(
                            label=f"📥 Скачать файл ({len(final_output)} уникальных городов)",
                            data=excel_bytes,
                            file_name=f"{safe_sheet_name}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary",
                            key=f"download_sheet_{sheet_name}_{tab_idx}"
                        )
                        
                        # Сохраняем в session_state для архива (обновляется при каждом изменении селектора)
                        st.session_state.vacancy_files[sheet_name] = {
                            'data': excel_bytes,
                            'name': f"{safe_sheet_name}.xlsx",
                            'count': len(final_output)
                        }
                    else:
                        st.warning("⚠️ Нет данных для выгрузки")
                        # Удаляем из vacancy_files если данных больше нет
                        if sheet_name in st.session_state.vacancy_files:
                            del st.session_state.vacancy_files[sheet_name]
                    
                    # Кнопка для скачивания всех файлов архивом
                    st.markdown("---")
                    st.markdown("### 📦 Скачать все вкладки одним архивом")
                    
                    if 'vacancy_files' in st.session_state and st.session_state.vacancy_files:
                        total_cities = sum(f['count'] for f in st.session_state.vacancy_files.values())
                        
                        if st.button("📦 Сформировать архив", use_container_width=True, type="primary", key="create_sheets_archive"):
                            zip_buffer = io.BytesIO()
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                for sheet_name, file_info in st.session_state.vacancy_files.items():
                                    zip_file.writestr(file_info['name'], file_info['data'])
                            
                            zip_buffer.seek(0)
                            
                            st.download_button(
                                label=f"📥 Скачать архив ({len(st.session_state.vacancy_files)} вкладок, {total_cities} городов)",
                                data=zip_buffer,
                                file_name=f"all_sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                mime="application/zip",
                                use_container_width=True,
                                type="secondary"
                            )
                    
                    # Останавливаем выполнение
                    st.stop()
                
                elif st.session_state.sheet_mode == 'columns':
                    # РЕЖИМ СТОЛБЦА: оригинальная логика с столбцом "Вакансия"
                    st.subheader("🎯 Редактирование и выгрузка по вакансиям")
                
                # Получаем названия столбцов
                original_cols = st.session_state.original_df.columns.tolist()
                
                # Находим столбец "Вакансия"
                vacancy_col = None
                for col in original_cols:
                    if 'вакансия' in str(col).lower():
                        vacancy_col = col
                        break
                
                if vacancy_col:
                    # FIX: Формируем данные для экспорта (исключаем не найденные с эмодзи)
                    export_df = result_df[
                        (result_df['Итоговое гео'].notna()) &
                        (~result_df['Статус'].str.contains('❌ Не найдено', na=False)) &
                        (~result_df['Статус'].str.contains('Пустое значение', na=False))
                    ].copy()

                    # КРИТИЧНО: Также исключаем ВСЕ дубликаты городов с "❌ Не найдено"
                    excluded_cities = result_df[
                        result_df['Статус'].str.contains('❌ Не найдено', na=False)
                    ]['Исходное название'].unique()

                    if len(excluded_cities) > 0:
                        excluded_normalized = set()
                        for city in excluded_cities:
                            if pd.notna(city):
                                normalized = str(city).replace('ё', 'е').replace('Ё', 'Е').lower().strip()
                                normalized = ' '.join(normalized.split())
                                excluded_normalized.add(normalized)

                        export_df['_temp_normalized'] = (
                            export_df['Исходное название']
                            .fillna('').astype(str)
                            .str.replace('ё', 'е').str.replace('Ё', 'Е')
                            .str.lower().str.strip()
                            .str.replace(r'\s+', ' ', regex=True)
                        )
                        export_df = export_df[~export_df['_temp_normalized'].isin(excluded_normalized)].copy()
                        export_df = export_df.drop(columns=['_temp_normalized'])

                    # Получаем уникальные вакансии
                    if vacancy_col in export_df.columns:
                        unique_vacancies = sorted(export_df[vacancy_col].dropna().unique())

                        # Инициализируем состояние для редактирования вакансий
                        if 'vacancy_edits' not in st.session_state:
                            st.session_state.vacancy_edits = {}

                        # Инициализируем выбранную вакансию
                        if 'selected_vacancy' not in st.session_state:
                            st.session_state.selected_vacancy = unique_vacancies[0]

                        # Выбор вакансии через radio кнопки
                        st.markdown("#### 📋 Выберите вакансию для редактирования:")
                        selected_vacancy = st.radio(
                            "Вакансия:",
                            options=unique_vacancies,
                            index=unique_vacancies.index(st.session_state.selected_vacancy) if st.session_state.selected_vacancy in unique_vacancies else 0,
                            key="vacancy_selector",
                            horizontal=False,
                            label_visibility="collapsed"
                        )
                        st.session_state.selected_vacancy = selected_vacancy

                        st.markdown("---")

                        # Показываем контент только для выбранной вакансии
                        vacancy = selected_vacancy
                        tab_idx = unique_vacancies.index(vacancy)

                        # Фильтруем данные по вакансии
                        vacancy_df = export_df[export_df[vacancy_col] == vacancy].copy()

                        # Показываем таблицу с возможностью редактирования
                        st.markdown("#### Города для редактирования (совпадение ≤ 95%)")

                        editable_vacancy_rows = vacancy_df[vacancy_df['Совпадение %'] <= 95].copy()
                        
                        # Убираем дубликаты по исходному названию для редактирования
                        if len(editable_vacancy_rows) > 0:
                            # VECTORIZED: normalize city name
                            editable_vacancy_rows['_normalized_original'] = (
                                editable_vacancy_rows['Исходное название']
                                .fillna('').astype(str)
                                .str.replace('ё', 'е').str.replace('Ё', 'Е')
                                .str.lower().str.strip()
                                .str.replace(r'\s+', ' ', regex=True)
                            )
                            editable_vacancy_rows = editable_vacancy_rows.drop_duplicates(subset=['_normalized_original'], keep='first')

                            # Сортируем: сначала "Нет совпадения", затем по возрастанию процента
                            # VECTORIZED: sort priority (0 for not found, 1 for others)
                            editable_vacancy_rows['_sort_priority'] = (~editable_vacancy_rows['Статус'].str.contains('❌ Не найдено', na=False)).astype(int)
                            editable_vacancy_rows = editable_vacancy_rows.sort_values(
                                ['_sort_priority', 'Совпадение %'],
                                ascending=[True, True]
                            )
                            editable_vacancy_rows = editable_vacancy_rows.drop(columns=['_sort_priority'])

                        if len(editable_vacancy_rows) > 0:

                            # Callback для сохранения выбора ТОЛЬКО при изменении
                            def on_city_select_vacancy(selection_key, widget_key):
                                """Callback для режима split vacancy - вызывается только при изменении"""
                                selected = st.session_state.get(widget_key)
                                if selected == "❌ Нет совпадения":
                                    st.session_state.manual_selections[selection_key] = "❌ Нет совпадения"
                                elif selected:
                                    # Извлекаем название без процента
                                    if "(" in selected and not selected.startswith("❌"):
                                        city_match = selected.rsplit(' (', 1)[0]
                                        st.session_state.manual_selections[selection_key] = city_match
                                    else:
                                        st.session_state.manual_selections[selection_key] = selected

                            # Получаем список всех городов России для выбора
                            russia_cities_for_select = []
                            for city_name, city_info in hh_areas.items():
                                if city_info.get('root_parent_id') == '113':
                                    russia_cities_for_select.append(city_name)
                            russia_cities_for_select = sorted(russia_cities_for_select)

                            for idx, row in editable_vacancy_rows.iterrows():
                                col1, col2, col3 = st.columns([2, 3, 1])
                                
                                with col1:
                                    st.markdown(f"**{row['Исходное название']}**")
                                
                                with col2:
                                    row_id = row['row_id']
                                    city_name = row['Исходное название']
                                    current_value = row['Итоговое гео']
                                    current_match = row['Совпадение %']

                                    # Используем кэш кандидатов из smart_match_city
                                    candidates = st.session_state.candidates_cache.get(row_id, [])

                                    # Если кэша нет, ищем заново (для обратной совместимости)
                                    if not candidates:
                                        # Используем только российские города
                                        # OPTIMIZED: use cached version
                                        candidates = get_candidates_by_word(city_name, get_russian_cities_cached(hh_areas), limit=20)
                                    
                                    # Если есть текущее значение из сопоставления - добавляем его в список
                                    if current_value and current_value != city_name:
                                        # Проверяем, есть ли уже это значение в кандидатах
                                        candidate_names = [c[0] for c in candidates]
                                        if current_value not in candidate_names:
                                            # Добавляем текущее значение в список
                                            candidates.append((current_value, current_match))

                                    # Сортируем кандидатов по убыванию процента совпадения
                                    candidates.sort(key=lambda x: x[1], reverse=True)

                                    # Формируем опции - всегда показываем топ кандидатов
                                    if candidates:
                                        options = ["❌ Нет совпадения"] + [f"{c[0]} ({c[1]:.1f}%)" for c in candidates[:20]]
                                    else:
                                        # Если совсем нет кандидатов - показываем хотя бы "Нет совпадения"
                                        options = ["❌ Нет совпадения"]
                                    
                                    # Уникальный ключ для каждой вакансии
                                    unique_key = f"select_{vacancy}_{row_id}_{tab_idx}"
                                    selection_key = (vacancy, row_id)

                                    if selection_key in st.session_state.manual_selections:
                                        selected_value = st.session_state.manual_selections[selection_key]
                                        if selected_value == "❌ Нет совпадения":
                                            default_idx = 0
                                        else:
                                            # Ищем в options, может быть как с процентом, так и без
                                            default_idx = 0
                                            for i, opt in enumerate(options):
                                                if selected_value in opt or opt.startswith(selected_value):
                                                    default_idx = i
                                                    break
                                    else:
                                        # Если manual_selections нет, используем current_value из результата сопоставления
                                        default_idx = 0
                                        if current_value:
                                            # Ищем current_value в options (может быть с процентом)
                                            for i, opt in enumerate(options):
                                                # opt вида "Город (Область) (90.0%)", current_value вида "Город (Область)"
                                                if opt.startswith(current_value) or current_value in opt:
                                                    default_idx = i
                                                    break
                                    
                                    st.selectbox(
                                        "Выберите город:",
                                        options=options,
                                        index=default_idx,
                                        key=unique_key,
                                        label_visibility="collapsed",
                                        on_change=on_city_select_vacancy,
                                        args=(selection_key, unique_key)
                                    )
                                
                                with col3:
                                    st.text(f"{row['Совпадение %']}%")
                                
                                st.markdown("<hr style='margin-top: 5px; margin-bottom: 5px;'>", unsafe_allow_html=True)

                        else:
                            st.success("✅ Все города распознаны корректно!")
                        
                        # ============================================
                        # БЛОК: ДОБАВЛЕНИЕ ГОРОДОВ ДЛЯ ЭТОЙ ВАКАНСИИ
                        # ============================================
                        st.markdown("---")
                        st.markdown("#### ➕ Добавить дополнительные города")
                        
                        # Инициализируем список добавленных городов для каждой вакансии
                        vacancy_key = f"added_cities_{vacancy}"
                        if vacancy_key not in st.session_state:
                            st.session_state[vacancy_key] = []
                        
                        # Селектор на половину ширины экрана
                        col_add_selector = st.columns([1, 1])
                        with col_add_selector[0]:
                            # Получаем только города России
                            russia_cities = []
                            for city_name, city_info in hh_areas.items():
                                if city_info.get('root_parent_id') == '113':
                                    russia_cities.append(city_name)

                            selected_add_city = st.selectbox(
                                "Выберите город:",
                                options=sorted(russia_cities),
                                key=f"city_selector_{vacancy}_{tab_idx}",
                                help="Выберите город из справочника HH.ru"
                            )

                        # Кнопки под селектором
                        col_add_btn1, col_add_btn2 = st.columns(2)
                        with col_add_btn1:
                            if st.button("➕ Добавить", use_container_width=True, type="secondary", key=f"add_btn_{vacancy}_{tab_idx}"):
                                if selected_add_city and selected_add_city not in st.session_state[vacancy_key]:
                                    # Безопасное добавление с проверкой лимитов
                                    if safe_session_append(vacancy_key, selected_add_city):
                                        st.success(f"✅ {selected_add_city}")
                                    else:
                                        st.error("⚠️ Достигнут лимит добавленных городов")
                                        log_security_event('session_limit', f"Limit reached for {vacancy_key}", 'WARNING')
                                elif selected_add_city in st.session_state[vacancy_key]:
                                    st.warning(f"⚠️ Уже добавлен")

                        with col_add_btn2:
                            if st.button("🗑️ Очистить", use_container_width=True, key=f"clear_btn_{vacancy}_{tab_idx}"):
                                st.session_state[vacancy_key] = []
                                st.rerun()
                        
                        # Показываем список добавленных городов
                        if st.session_state[vacancy_key]:
                            st.info(f"📋 Добавлено городов: **{len(st.session_state[vacancy_key])}**")
                            added_text = ", ".join(st.session_state[vacancy_key])
                            st.text_area(
                                "Список:",
                                value=added_text,
                                height=80,
                                disabled=True,
                                label_visibility="collapsed",
                                key=f"added_list_{vacancy}_{tab_idx}"
                            )
                        
                        st.markdown("---")
                        
                        # Применяем ручные изменения через КЭШИРОВАННУЮ функцию
                        # Фильтруем только изменения для текущей вакансии
                        vacancy_selections = {}
                        for selection_key, new_value in st.session_state.manual_selections.items():
                            if isinstance(selection_key, tuple):
                                key_vacancy, row_id = selection_key
                                if key_vacancy == vacancy:
                                    vacancy_selections[row_id] = new_value
                            else:
                                # Для обратной совместимости
                                vacancy_selections[selection_key] = new_value

                        # Используем кэшированную функцию вместо цикла
                        # FIX: Передаем vacancy в cache_key для уникальности кэша каждой вакансии
                        vacancy_final_df = apply_manual_selections_cached(
                            vacancy_df,
                            vacancy_selections,
                            hh_areas,
                            cache_key=f"vacancy_{vacancy}"
                        )

                        # КРИТИЧНЫЙ FIX: Применяем изменения ко ВСЕМ дубликатам
                        # Проблема: если в файле 2 строки "Москва", показывается только 1 в редактировании
                        # При изменении на "Питер", только 1 строка меняется, вторая остается "Москва"
                        # Решение: найти все строки с таким же исходным названием и применить то же изменение
                        for row_id_changed, new_value in vacancy_selections.items():
                            # Находим измененную строку
                            changed_row = vacancy_final_df[vacancy_final_df['row_id'] == row_id_changed]
                            if len(changed_row) == 0:
                                continue

                            # Получаем исходное название этой строки
                            original_city = changed_row['Исходное название'].values[0]

                            # Нормализуем для поиска дубликатов
                            original_normalized = str(original_city).replace('ё', 'е').replace('Ё', 'Е').lower().strip()
                            original_normalized = ' '.join(original_normalized.split())

                            # Находим ВСЕ строки с таким же исходным названием
                            vacancy_final_df['_temp_norm'] = (
                                vacancy_final_df['Исходное название']
                                .fillna('').astype(str)
                                .str.replace('ё', 'е').str.replace('Ё', 'Е')
                                .str.lower().str.strip()
                                .str.replace(r'\s+', ' ', regex=True)
                            )
                            duplicate_mask = (vacancy_final_df['_temp_norm'] == original_normalized)
                            vacancy_final_df = vacancy_final_df.drop(columns=['_temp_norm'])

                            # Применяем то же изменение ко ВСЕМ дубликатам
                            if new_value == "❌ Нет совпадения":
                                vacancy_final_df.loc[duplicate_mask, 'Итоговое гео'] = None
                                vacancy_final_df.loc[duplicate_mask, 'ID HH'] = None
                                vacancy_final_df.loc[duplicate_mask, 'Регион'] = None
                                vacancy_final_df.loc[duplicate_mask, 'Совпадение %'] = 0
                                vacancy_final_df.loc[duplicate_mask, 'Изменение'] = 'Нет'
                                vacancy_final_df.loc[duplicate_mask, 'Статус'] = '❌ Не найдено'
                            else:
                                vacancy_final_df.loc[duplicate_mask, 'Итоговое гео'] = new_value
                                if new_value in hh_areas:
                                    vacancy_final_df.loc[duplicate_mask, 'ID HH'] = hh_areas[new_value]['id']
                                    vacancy_final_df.loc[duplicate_mask, 'Регион'] = hh_areas[new_value]['parent']
                                vacancy_final_df.loc[duplicate_mask, 'Изменение'] = 'Да'

                        # FIX: Исключаем не найденные (❌ Не найдено) для публикатора
                        temp_vacancy_df = vacancy_final_df[
                            (vacancy_final_df['Итоговое гео'].notna()) &
                            (~vacancy_final_df['Статус'].str.contains('❌ Не найдено', na=False))
                        ].copy()

                        # КРИТИЧНО: Также исключаем ВСЕ дубликаты городов с "❌ Не найдено"
                        excluded_cities = vacancy_final_df[
                            vacancy_final_df['Статус'].str.contains('❌ Не найдено', na=False)
                        ]['Исходное название'].unique()

                        if len(excluded_cities) > 0:
                            excluded_normalized = set()
                            for city in excluded_cities:
                                if pd.notna(city):
                                    normalized = str(city).replace('ё', 'е').replace('Ё', 'Е').lower().strip()
                                    normalized = ' '.join(normalized.split())
                                    excluded_normalized.add(normalized)

                            temp_vacancy_df['_temp_normalized'] = (
                                temp_vacancy_df['Исходное название']
                                .fillna('').astype(str)
                                .str.replace('ё', 'е').str.replace('Ё', 'Е')
                                .str.lower().str.strip()
                                .str.replace(r'\s+', ' ', regex=True)
                            )
                            temp_vacancy_df = temp_vacancy_df[~temp_vacancy_df['_temp_normalized'].isin(excluded_normalized)].copy()
                            temp_vacancy_df = temp_vacancy_df.drop(columns=['_temp_normalized'])

                        vacancy_final_df = temp_vacancy_df

                        # Формируем DataFrame для выгрузки
                        output_vacancy_df = pd.DataFrame()
                        output_vacancy_df[original_cols[0]] = vacancy_final_df['Итоговое гео']
                        
                        for col in original_cols[1:]:
                            if col != vacancy_col and col in vacancy_final_df.columns:
                                output_vacancy_df[col] = vacancy_final_df[col].values
                        
                        # Добавляем дополнительные города для этой вакансии
                        vacancy_key = f"added_cities_{vacancy}"
                        if vacancy_key in st.session_state and st.session_state[vacancy_key]:
                            # Получаем последнюю строку для значений других столбцов
                            if len(output_vacancy_df) > 0:
                                last_row_values = output_vacancy_df.iloc[-1].tolist()
                                
                                for add_city in st.session_state[vacancy_key]:
                                    new_row = [add_city] + last_row_values[1:]
                                    output_vacancy_df.loc[len(output_vacancy_df)] = new_row
                        
                        # Удаляем дубликаты по городу
                        # VECTORIZED: normalize city name
                        output_vacancy_df['_normalized'] = (
                            output_vacancy_df[original_cols[0]]
                            .fillna('').astype(str)
                            .str.replace('ё', 'е').str.replace('Ё', 'Е')
                            .str.lower().str.strip()
                            .str.replace(r'\s+', ' ', regex=True)
                        )
                        output_vacancy_df = output_vacancy_df.drop_duplicates(subset=['_normalized'], keep='first')
                        output_vacancy_df = output_vacancy_df.drop(columns=['_normalized'])

                        # Удаляем первую строку, если она является заголовком
                        output_vacancy_df = remove_header_row_if_needed(output_vacancy_df, original_cols[0])

                        # Проверяем что есть данные для выгрузки
                        if len(output_vacancy_df) > 0:
                            # Превью итогового файла для вакансии
                            st.markdown(f"#### 👀 Превью итогового файла - {vacancy}")
                            st.dataframe(output_vacancy_df, use_container_width=True, height=300)

                            # Кнопка выгрузки для этой вакансии
                            st.markdown("---")
                            safe_vacancy_name = str(vacancy).replace('/', '_').replace('\\', '_')[:50]

                            # Санитизация данных перед экспортом (защита от CSV Injection)
                            output_vacancy_df = sanitize_csv_content(output_vacancy_df)

                            # OPTIMIZED: Используем кэшированную генерацию файла
                            excel_bytes = create_excel_bytes_cached(output_vacancy_df, 'Результат')
                            
                            st.download_button(
                                label=f"📥 Скачать файл ({len(output_vacancy_df)} уникальных городов)",
                                data=excel_bytes,
                                file_name=f"{safe_vacancy_name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="primary",
                                key=f"download_{vacancy}_{tab_idx}"
                            )

                            # Сохраняем файл в session_state для архива (обновляется при каждом изменении селектора)
                            if 'vacancy_files' not in st.session_state:
                                st.session_state.vacancy_files = {}
                            st.session_state.vacancy_files[vacancy] = {
                                'data': excel_bytes,
                                'name': f"{safe_vacancy_name}.xlsx",
                                'count': len(output_vacancy_df)
                            }
                        else:
                            st.warning("⚠️ Нет данных для выгрузки")
                            # Удаляем из vacancy_files если данных больше нет
                            if 'vacancy_files' in st.session_state and vacancy in st.session_state.vacancy_files:
                                del st.session_state.vacancy_files[vacancy]
                        
                        # Кнопка для скачивания всех файлов архивом
                        st.markdown("---")
                        st.markdown("### 📦 Скачать все вакансии одним архивом")
                        
                        # Проверяем что есть сохраненные файлы
                        if 'vacancy_files' in st.session_state and st.session_state.vacancy_files:
                            total_cities = sum(f['count'] for f in st.session_state.vacancy_files.values())
                            
                            if st.button("📦 Сформировать архив", use_container_width=True, type="primary"):
                                # Создаем ZIP-архив из сохраненных файлов
                                zip_buffer = io.BytesIO()
                                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                    for vacancy_name, file_info in st.session_state.vacancy_files.items():
                                        zip_file.writestr(file_info['name'], file_info['data'])
                                
                                zip_buffer.seek(0)
                                
                                st.download_button(
                                    label=f"📥 Скачать архив ({len(st.session_state.vacancy_files)} вакансий, {total_cities} городов)",
                                    data=zip_buffer,
                                    file_name=f"all_vacancies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                    mime="application/zip",
                                    use_container_width=True,
                                    type="secondary"
                                )
                        else:
                            st.info("ℹ️ Пройдитесь по всем вкладкам, чтобы сформировать архив")
                
                # После обработки всех вакансий - останавливаем выполнение
                # Не показываем стандартные блоки для режима split
                st.stop()

            else:
                # ОБЫЧНЫЙ РЕЖИМ (как было раньше)
                col1, col2 = st.columns(2)
                
                with col1:
                    # Формируем файл для публикатора с исходными столбцами
                    # FIX: Исключаем не найденные (❌ Не найдено) и дубликаты
                    export_df = final_result_df[
                        (~final_result_df['Статус'].str.contains('Дубликат', na=False)) &
                        (final_result_df['Итоговое гео'].notna()) &
                        (~final_result_df['Статус'].str.contains('❌ Не найдено', na=False))
                    ].copy()

                    # КРИТИЧНО: Также исключаем ВСЕ дубликаты городов с "❌ Не найдено"
                    excluded_cities = final_result_df[
                        final_result_df['Статус'].str.contains('❌ Не найдено', na=False)
                    ]['Исходное название'].unique()

                    if len(excluded_cities) > 0:
                        excluded_normalized = set()
                        for city in excluded_cities:
                            if pd.notna(city):
                                normalized = str(city).replace('ё', 'е').replace('Ё', 'Е').lower().strip()
                                normalized = ' '.join(normalized.split())
                                excluded_normalized.add(normalized)

                        export_df['_temp_normalized'] = (
                            export_df['Исходное название']
                            .fillna('').astype(str)
                            .str.replace('ё', 'е').str.replace('Ё', 'Е')
                            .str.lower().str.strip()
                            .str.replace(r'\s+', ' ', regex=True)
                        )
                        export_df = export_df[~export_df['_temp_normalized'].isin(excluded_normalized)].copy()
                        export_df = export_df.drop(columns=['_temp_normalized'])

                    # Получаем названия столбцов из исходного файла
                    original_cols = st.session_state.original_df.columns.tolist()
                    
                    # Формируем итоговый DataFrame: первый столбец - итоговое гео, остальные - из исходного файла
                    publisher_df = pd.DataFrame()
                    publisher_df[original_cols[0]] = export_df['Итоговое гео']
                    
                    # Добавляем остальные столбцы из исходного файла
                    for col in original_cols[1:]:
                        if col in export_df.columns:
                            publisher_df[col] = export_df[col].values

                    # Удаляем первую строку, если она является заголовком (применяется до добавления городов)
                    publisher_df = remove_header_row_if_needed(publisher_df, original_cols[0])

                    # Добавляем дополнительные города с значениями из последней строки
                    if st.session_state.added_cities:
                        # Получаем последнюю строку из исходного файла
                        last_row_values = st.session_state.original_df.iloc[-1].tolist()
                        
                        for city in st.session_state.added_cities:
                            new_row = [city] + last_row_values[1:]  # Город + остальные значения из последней строки
                            publisher_df.loc[len(publisher_df)] = new_row
                        
                        # Удаляем дубликаты
                        # VECTORIZED: normalize city name
                        publisher_df['_normalized'] = (
                            publisher_df[original_cols[0]]
                            .fillna('').astype(str)
                            .str.replace('ё', 'е').str.replace('Ё', 'Е')
                            .str.lower().str.strip()
                            .str.replace(r'\s+', ' ', regex=True)
                        )
                        publisher_df = publisher_df.drop_duplicates(subset=['_normalized'], keep='first')
                        publisher_df = publisher_df.drop(columns=['_normalized'])

                    # Санитизация данных перед экспортом (защита от CSV Injection)
                    publisher_df = sanitize_csv_content(publisher_df)

                    output_publisher = io.BytesIO()
                    with pd.ExcelWriter(output_publisher, engine='openpyxl') as writer:
                        publisher_df.to_excel(writer, index=False, header=False, sheet_name='Результат')  
                    output_publisher.seek(0)  
                      
                    publisher_count = len(publisher_df)  
                      
                    st.download_button(
                        label=f"📤 Файл для публикатора\n{publisher_count} строк",
                        data=output_publisher,
                        file_name=f"geo_result_{uploaded_file.name.rsplit('.', 1)[0]}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key='download_publisher'
                    )

                with col2:
                    export_full_df = final_result_df.drop(['row_id', 'sort_priority'], axis=1, errors='ignore')

                    # Санитизация данных перед экспортом (защита от CSV Injection)
                    export_full_df = sanitize_csv_content(export_full_df)

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        export_full_df.to_excel(writer, index=False, sheet_name='Результат')  
                    output.seek(0)  
                      
                    st.download_button(
                        label="📥 Полный отчет с анализом",
                        data=output,
                        file_name=f"full_report_{uploaded_file.name.rsplit('.', 1)[0]}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key='download_full'
                    )
      
    except Exception as e:
        st.error(f"❌ Ошибка обработки файла: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
else:
    # Файл был удален - очищаем все кэши
    if 'processed' in st.session_state and st.session_state.processed:
        # Очищаем все кэши связанные с обработкой файла
        st.session_state.processed = False
        st.session_state.result_df = None
        st.session_state.manual_selections = {}
        st.session_state.added_cities = []
        st.session_state.candidates_cache = {}
        st.session_state.original_df = None
        if 'vacancy_files' in st.session_state:
            del st.session_state.vacancy_files
        if 'sheets_results' in st.session_state:
            del st.session_state.sheets_results
        # Очищаем кэши пагинации
        keys_to_delete = [k for k in st.session_state.keys() if k.startswith('edit_page')]
        for key in keys_to_delete:
            del st.session_state[key]

# ============================================
# БЛОК: ВЫБОР РЕГИОНОВ И ГОРОДОВ
# ============================================
st.markdown('<div id="выбор-регионов-и-городов"></div>', unsafe_allow_html=True)

# CSS для ЧЕРНОЙ окантовки multiselect (базовый цвет как у selectbox)
st.markdown("""
<style>
/* Черная окантовка для multiselect */
[data-testid="stMultiSelect"] div[data-baseweb="select"] > div {
    border-color: #1a1a1a !important;
}
[data-testid="stMultiSelect"] div[data-baseweb="select"] > div:hover {
    border-color: #1a1a1a !important;
}
[data-testid="stMultiSelect"] div[data-baseweb="select"] > div:focus-within {
    border-color: #1a1a1a !important;
    box-shadow: 0 0 0 0.2rem rgba(26, 26, 26, 0.25) !important;
}
</style>
""", unsafe_allow_html=True)

st.header("🗺️ Выбор регионов и городов")

if hh_areas is not None:
    # Получаем полный список городов для фильтров
    all_cities_full = get_all_cities(hh_areas)

    # ФИЛЬТРЫ В ОДНОМ БЛОКЕ
    st.markdown("### 🔍 Фильтры")
    col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)

    with col_filter1:
        # Форматируем федеральные округа с количеством регионов
        districts_formatted = []
        districts_mapping = {}
        for district, regions in FEDERAL_DISTRICTS.items():
            formatted = f"{district} ({len(regions)} рег.)"
            districts_formatted.append(formatted)
            districts_mapping[formatted] = district

        selected_districts_formatted = st.multiselect(
            "Федеральные округа:",
            options=districts_formatted,
            help="Можно выбрать несколько",
            key="districts_select"
        )

        # Получаем оригинальные названия округов
        selected_districts = [districts_mapping[d] for d in selected_districts_formatted]

    # Формируем список доступных регионов на основе выбранных округов
    available_regions = []
    if selected_districts:
        for district in selected_districts:
            available_regions.extend(FEDERAL_DISTRICTS[district])
    else:
        for regions in FEDERAL_DISTRICTS.values():
            available_regions.extend(regions)

    with col_filter2:
        # Форматируем регионы с указанием федерального округа
        regions_formatted = []
        regions_mapping = {}
        for region in sorted(available_regions):
            # Находим федеральный округ для региона
            fed_district = get_federal_district_by_region(region)
            if fed_district != "Не определен":
                # Сокращаем название округа для компактности
                district_short = fed_district.replace("Федеральный округ", "ФО").replace("федеральный округ", "ФО")
                formatted = f"{region} ({district_short})"
            else:
                formatted = region
            regions_formatted.append(formatted)
            regions_mapping[formatted] = region

        selected_regions_formatted = st.multiselect(
            "Области/Регионы:",
            options=regions_formatted,
            help="Можно выбрать несколько",
            key="regions_select"
        )

        # Получаем оригинальные названия регионов
        selected_regions = [regions_mapping[r] for r in selected_regions_formatted]

    with col_filter3:
        # Фильтр по часовому поясу (мультиселект)
        if not all_cities_full.empty:
            unique_timezones = sorted([tz for tz in all_cities_full['UTC'].unique() if tz and str(tz) != 'nan'])

            # Создаем форматированные опции с разницей от МСК
            timezone_options_formatted = []
            timezone_mapping = {}  # Маппинг отформатированных значений на оригинальные UTC

            for tz in unique_timezones:
                try:
                    # Парсим UTC offset для вычисления разницы с Москвой
                    sign = 1 if tz[0] == '+' else -1
                    hours = int(tz[1:3])
                    tz_hours = sign * hours
                    diff_msk = tz_hours - 3  # Москва = UTC+3

                    if diff_msk == 0:
                        formatted = f"{tz} (МСК)"
                    elif diff_msk > 0:
                        formatted = f"{tz} (+{diff_msk}ч от МСК)"
                    else:
                        formatted = f"{tz} ({diff_msk}ч от МСК)"

                    timezone_options_formatted.append(formatted)
                    timezone_mapping[formatted] = tz
                except (ValueError, IndexError, TypeError) as e:
                    # Если не удалось распарсить, добавляем как есть
                    logger.warning(f"Не удалось распарсить timezone '{tz}': {e}")
                    timezone_options_formatted.append(tz)
                    timezone_mapping[tz] = tz

            selected_timezones_formatted = st.multiselect(
                "Часовой пояс (UTC):",
                options=timezone_options_formatted,
                help="Можно выбрать несколько",
                key="timezone_filter"
            )

            # Получаем оригинальные значения UTC
            selected_timezones = [timezone_mapping.get(tz_fmt, tz_fmt) for tz_fmt in selected_timezones_formatted]
        else:
            selected_timezones = []

    with col_filter4:
        # Выбор городов (множественный выбор)
        if not all_cities_full.empty:
            city_options = sorted(all_cities_full['Город'].unique())
            selected_cities = st.multiselect(
                "Выбрать город:",
                options=city_options,
                help="Можно выбрать несколько",
                key="cities_multiselect"
            )
        else:
            selected_cities = []

    # ВТОРАЯ СТРОКА ФИЛЬТРОВ - Население
    st.markdown("---")
    col_filter_pop1, col_filter_pop2 = st.columns([1, 3])

    with col_filter_pop1:
        # Фильтр по населению (multiselect)
        if not all_cities_full.empty and 'Население' in all_cities_full.columns:
            # Определяем диапазоны населения
            population_ranges = {
                "До 10,000 человек": (1, 10_000),
                "10,000 - 100,000 человек": (10_000, 100_000),
                "100,000 - 500,000 человек": (100_000, 500_000),
                "500,000 - 1,000,000 человек": (500_000, 1_000_000),
                "Более 1,000,000 человек": (1_000_000, float('inf'))
            }

            selected_population_ranges = st.multiselect(
                "Население (жители):",
                options=list(population_ranges.keys()),
                help="Можно выбрать несколько диапазонов",
                key="population_filter"
            )
        else:
            selected_population_ranges = []
            population_ranges = {}

    # Определяем, какие регионы использовать для поиска
    regions_to_search = []
    if selected_regions:
        regions_to_search = selected_regions
    elif selected_districts:
        for district in selected_districts:
            regions_to_search.extend(FEDERAL_DISTRICTS[district])

    # Очищаем превью если все фильтры сняты
    if not regions_to_search and not selected_cities and not selected_timezones and not selected_population_ranges:
        if 'regions_cities_df' in st.session_state:
            del st.session_state.regions_cities_df

    # Функция для фильтрации по населению
    def filter_by_population(df, selected_ranges, ranges_dict):
        """Фильтрует DataFrame по выбранным диапазонам населения"""
        if not selected_ranges or df.empty or 'Население' not in df.columns:
            return df

        # Создаем маску для фильтрации
        mask = pd.Series([False] * len(df), index=df.index)
        for range_name in selected_ranges:
            min_pop, max_pop = ranges_dict[range_name]
            mask |= (df['Население'] >= min_pop) & (df['Население'] < max_pop)

        return df[mask]

    # КНОПКИ ДЕЙСТВИЙ
    col_btn1, col_btn2, col_btn3 = st.columns(3)

    with col_btn1:
        # Показываем кнопку только если что-то выбрано
        if regions_to_search:
            # Информация о выборе
            if selected_regions:
                st.info(f"📍 Выбрано регионов: **{len(selected_regions)}**")
            elif selected_districts:
                st.info(f"📍 Выбрано округов: **{len(selected_districts)}** (включает {len(regions_to_search)} регионов)")

            if st.button("🔍 Получить список городов по регионам", type="primary", use_container_width=True):
                with st.spinner("Формирую список городов..."):
                    # Очищаем старые результаты
                    if 'city_df' in st.session_state:
                        del st.session_state.city_df
                    if 'timezones_df' in st.session_state:
                        del st.session_state.timezones_df
                    # Получаем список городов по регионам
                    result_df = get_cities_by_regions(hh_areas, regions_to_search)
                    # Применяем фильтр по населению
                    result_df = filter_by_population(result_df, selected_population_ranges, population_ranges)
                    # Сохраняем новый результат
                    st.session_state.regions_cities_df = result_df

    with col_btn2:
        # Кнопка для выбранных городов
        if selected_cities:
            # Информация о выборе
            if len(selected_cities) == 1:
                st.info(f"📍 Выбран город: **{selected_cities[0]}**")
            else:
                st.info(f"📍 Выбрано городов: **{len(selected_cities)}**")

            if st.button(f"🔍 Получить информацию о {'городе' if len(selected_cities) == 1 else 'городах'}", type="primary", use_container_width=True):
                with st.spinner(f"Получаю информацию о {len(selected_cities)} {'городе' if len(selected_cities) == 1 else 'городах'}..."):
                    # Очищаем старые результаты
                    if 'city_df' in st.session_state:
                        del st.session_state.city_df
                    if 'timezones_df' in st.session_state:
                        del st.session_state.timezones_df
                    # Фильтруем данные по выбранным городам
                    city_df = all_cities_full[all_cities_full['Город'].isin(selected_cities)].copy()
                    # Применяем фильтр по населению
                    city_df = filter_by_population(city_df, selected_population_ranges, population_ranges)
                    # Сохраняем в общий результат
                    if not city_df.empty:
                        st.session_state.regions_cities_df = city_df
        # Кнопка для фильтра по населению (если выбрано только население)
        elif selected_population_ranges and not regions_to_search and not selected_timezones:
            # Информация о выборе
            if len(selected_population_ranges) == 1:
                st.info(f"👥 Выбран диапазон: **{selected_population_ranges[0]}**")
            else:
                st.info(f"👥 Выбрано диапазонов: **{len(selected_population_ranges)}**")

            if st.button("🔍 Получить список городов по населению", type="primary", use_container_width=True):
                with st.spinner("Фильтрую по населению..."):
                    # Очищаем старые результаты
                    if 'city_df' in st.session_state:
                        del st.session_state.city_df
                    if 'timezones_df' in st.session_state:
                        del st.session_state.timezones_df
                    # Берем все города и фильтруем по населению
                    result_df = all_cities_full.copy()
                    result_df = filter_by_population(result_df, selected_population_ranges, population_ranges)
                    # Сохраняем результат
                    if not result_df.empty:
                        st.session_state.regions_cities_df = result_df

    with col_btn3:
        # Кнопка для выгрузки по часовым поясам
        if selected_timezones:
            # Информация о выборе
            if len(selected_timezones) == 1:
                st.info(f"🕐 Выбран часовой пояс: **{selected_timezones[0]}**")
            else:
                st.info(f"🕐 Выбрано часовых поясов: **{len(selected_timezones)}**")

            # Формируем текст для кнопки
            if len(selected_timezones) == 1:
                button_text = f"🔍 Получить список городов по UTC"
            else:
                button_text = f"🔍 Получить список городов"

            if st.button(button_text, type="primary", use_container_width=True):
                with st.spinner(f"Фильтрую по выбранным часовым поясам..."):
                    # Очищаем старые результаты
                    if 'city_df' in st.session_state:
                        del st.session_state.city_df
                    if 'timezones_df' in st.session_state:
                        del st.session_state.timezones_df
                    # Фильтруем города по выбранным часовым поясам
                    filtered_df = all_cities_full[all_cities_full['UTC'].isin(selected_timezones)].copy()
                    # Применяем фильтр по населению
                    filtered_df = filter_by_population(filtered_df, selected_population_ranges, population_ranges)
                    # Сохраняем в общий результат
                    if not filtered_df.empty:
                        st.session_state.regions_cities_df = filtered_df

    # ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ (ТАБЛИЦА ПРЕВЬЮ НА ПОЛНУЮ ШИРИНУ)
    # Единый блок для отображения результатов всех фильтров
    if 'regions_cities_df' in st.session_state and not st.session_state.regions_cities_df.empty:
        cities_df = st.session_state.regions_cities_df

        # Определяем количество найденных городов
        city_count = len(cities_df)

        # Универсальное сообщение
        if city_count == 1:
            st.success(f"✅ Найден **{city_count}** город")
        else:
            st.success(f"✅ Найдено **{city_count}** городов")

        # Показываем таблицу на полную ширину
        # Сортируем по населению по убыванию
        display_cities_df = cities_df.copy()
        if 'Население' in display_cities_df.columns:
            display_cities_df = display_cities_df.sort_values('Население', ascending=False)
        display_cities_df = display_cities_df.reset_index(drop=True)

        st.dataframe(display_cities_df, use_container_width=True, height=400, hide_index=True)

        # Кнопки для скачивания
        col1, col2 = st.columns(2)

        with col1:
            # Полный отчет
            # Санитизация данных перед экспортом (защита от CSV Injection)
            sanitized_cities_df = sanitize_csv_content(cities_df)

            output_full = io.BytesIO()
            with pd.ExcelWriter(output_full, engine='openpyxl') as writer:
                sanitized_cities_df.to_excel(writer, index=False, sheet_name='Города')
            output_full.seek(0)

            st.download_button(
                label=f"📥 Скачать полный отчет ({city_count} городов)" if city_count > 1 else f"📥 Скачать полный отчет ({city_count} город)",
                data=output_full,
                file_name="cities_full_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_regions_full"
            )

        with col2:
            # Только названия городов для публикатора
            publisher_df = pd.DataFrame({'Город': cities_df['Город']})

            # Санитизация данных перед экспортом (защита от CSV Injection)
            publisher_df = sanitize_csv_content(publisher_df)

            output_publisher = io.BytesIO()
            with pd.ExcelWriter(output_publisher, engine='openpyxl') as writer:
                publisher_df.to_excel(writer, index=False, header=False, sheet_name='Гео')
            output_publisher.seek(0)

            st.download_button(
                label=f"📤 Для публикатора ({city_count} городов)" if city_count > 1 else f"📤 Для публикатора ({city_count} город)",
                data=output_publisher,
                file_name="cities_for_publisher.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_regions_publisher"
            )

# ============================================
# БЛОК: ОБЪЕДИНИТЕЛЬ ФАЙЛОВ
# ============================================
st.markdown('<div id="объединитель-файлов"></div>', unsafe_allow_html=True)
st.header("🔗 Объединитель файлов")

st.markdown("""
Загрузите несколько файлов с одинаковыми столбцами. Инструмент объединит их в один файл.
Полные дубликаты будут выделены оранжевым цветом и размещены вначале.
""")

uploaded_files = st.file_uploader(
    "Загрузите файлы для объединения",
    type=['xlsx', 'xls', 'xlsm', 'xlsb', 'csv'],
    accept_multiple_files=True,
    key="file_merger_uploader",
    help="Можно загрузить несколько файлов Excel (xlsx, xls, xlsm, xlsb) или CSV с одинаковыми столбцами"
)

if uploaded_files:
    # Валидация размера и расширения файлов
    files_valid = True
    for uploaded_file in uploaded_files:
        # Проверка размера
        is_valid_size, error_msg = validate_file_size(uploaded_file.size)
        if not is_valid_size:
            st.error(f"❌ {uploaded_file.name}: {error_msg}")
            logger.warning(f"Файл отклонен (размер): {uploaded_file.name} ({uploaded_file.size} байт)")
            log_security_event('file_size_exceeded', f"{uploaded_file.name}: {uploaded_file.size} байт", 'WARNING')
            files_valid = False

        # Проверка расширения
        is_valid_ext, error_msg = validate_file_extension(uploaded_file.name, ['.xlsx', '.xls', '.xlsm', '.xlsb', '.csv'])
        if not is_valid_ext:
            st.error(f"❌ {uploaded_file.name}: {error_msg}")
            logger.warning(f"Файл отклонен (расширение): {uploaded_file.name}")
            log_security_event('invalid_file_extension', uploaded_file.name, 'WARNING')
            files_valid = False

    if not files_valid:
        st.stop()

    try:
        with st.spinner("Обрабатываем файлы..."):
            # Читаем все файлы
            all_dataframes = []
            for uploaded_file in uploaded_files:
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)
                all_dataframes.append(df)
                st.success(f"✅ Загружен: {uploaded_file.name} ({len(df)} строк)")

            # Объединяем все файлы
            merged_df = pd.concat(all_dataframes, ignore_index=True)

            # Находим полные дубликаты
            duplicates_mask = merged_df.duplicated(keep=False)
            duplicates = merged_df[duplicates_mask].copy()
            non_duplicates = merged_df[~duplicates_mask].copy()

            # Создаем итоговый DataFrame: сначала дубликаты, затем остальные
            final_df = pd.concat([duplicates, non_duplicates], ignore_index=True)

            # Статистика
            total_rows = len(merged_df)
            duplicate_rows = len(duplicates)
            unique_rows = len(non_duplicates)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Всего строк", total_rows)
            with col2:
                st.metric("Дубликаты", duplicate_rows)
            with col3:
                st.metric("Уникальные", unique_rows)

            if duplicate_rows > 0:
                st.warning(f"⚠️ Найдено {duplicate_rows} дубликатов. Они будут выделены оранжевым цветом в скачанном файле и размещены вначале.")
            else:
                st.success("✅ Дубликаты не найдены!")

            st.markdown("### 👀 Превью объединенного файла")
            st.info(f"ℹ️ Первые {duplicate_rows} строк - дубликаты (будут выделены оранжевым в Excel)")
            st.dataframe(final_df, use_container_width=True, height=400)

            # Санитизация данных перед экспортом (защита от CSV Injection)
            sanitized_final_df = sanitize_csv_content(final_df)

            # Кнопка скачивания
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                sanitized_final_df.to_excel(writer, index=False, sheet_name='Объединенные данные')

                # Применяем оранжевый цвет к дубликатам в Excel
                workbook = writer.book
                worksheet = writer.sheets['Объединенные данные']

                from openpyxl.styles import PatternFill
                orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

                # Выделяем дубликаты (начиная со строки 2, т.к. строка 1 - заголовок)
                for row_idx in range(2, duplicate_rows + 2):
                    for col_idx in range(1, len(final_df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = orange_fill

            output.seek(0)

            st.download_button(
                label=f"📥 Скачать объединенный файл ({total_rows} строк)",
                data=output,
                file_name=f"merged_file_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_merged_file"
            )

    except Exception as e:
        st.error(f"❌ Ошибка при обработке файлов: {str(e)}")
        st.info("Убедитесь, что все файлы имеют одинаковые столбцы.")

st.markdown("---")

# =====================================================
# Раздел: Сверки с клиентами
# =====================================================
st.markdown('<div id="сверки-с-клиентами"></div>', unsafe_allow_html=True)

# CSS для селекторов в разделе Сверки с клиентами
st.markdown("""
<style>
/* Красная окантовка для selectbox и multiselect в разделе Сверки */
[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
    border-color: #e14531 !important;
}
[data-testid="stSelectbox"] div[data-baseweb="select"] > div:hover {
    border-color: #e14531 !important;
}
[data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus-within {
    border-color: #e14531 !important;
    box-shadow: 0 0 0 0.2rem rgba(225, 69, 49, 0.25) !important;
}
</style>
""", unsafe_allow_html=True)

st.header("🔄 Сверки с клиентами")

st.markdown("""
Инструменты для сверки данных с различными клиентами. Выберите нужную сверку ниже.
""")

# Открываем контейнер для стилизации блоков кода
st.markdown('<div class="matrix-code-section">', unsafe_allow_html=True)

# Яндекс.Еда - активная сверка
with st.expander("Яндекс.Еда", expanded=False):
    st.markdown("""
    ### Инструкция по запуску

    **ВАЖНО!** Скрипт нужно запускать в Google Colab в строгой последовательности:

    1. Откройте [Google Colab](https://colab.research.google.com/)
    2. Создайте **две отдельные ячейки**
    3. В первую ячейку скопируйте **Блок 1** (установка библиотек)
    4. Запустите первую ячейку и **дождитесь завершения установки**
    5. Во вторую ячейку скопируйте **Блок 2** (основной код)
    6. Запустите вторую ячейку
    7. Загрузите файлы когда появится запрос

    #### Ожидаемые файлы:
    - **"ООО Хэдхантер Биллинг....."** - отчет биллинг
    - **"Отчет-по-откликам-по-проектам-работодателя-"** - отчет внутренний hh
    - **"Leads_"** - лиды из ЛК Я.Еды
    """)

    st.markdown("---")

    # Блок 1: Установка библиотек
    st.markdown("### Блок 1: Установка библиотек")
    st.markdown("**Запустите этот код ПЕРВЫМ в отдельной ячейке Google Colab:**")

    libs_code = """!pip install pandas openpyxl fuzzywuzzy python-Levenshtein"""

    st.code(libs_code, language="python")

    st.markdown("---")

    # Блок 2: Основной код
    st.markdown("### Блок 2: Основной код сверки")
    st.markdown("**Запустите этот код ВТОРЫМ после установки библиотек:**")

    st.info("**Кнопка копирования** в правом верхнем углу блока кода скопирует **весь код целиком**")

    # Безопасное чтение основного кода из файла
    try:
        full_code = safe_read_file("yaedamatch", encoding="utf-8")

        if full_code:
            # Извлекаем код начиная с импорта библиотек
            main_code_start = full_code.find("# ============================================\n# ИМПОРТ БИБЛИОТЕК")
            if main_code_start != -1:
                main_code = full_code[main_code_start:]
            else:
                main_code = full_code

            # Отображаем код
            with st.expander("Показать код", expanded=False):
                st.code(main_code, language="python", line_numbers=False)
        else:
            st.error("❌ Не удалось загрузить файл yaedamatch")

    except FileNotFoundError:
        st.error("Файл yaedamatch не найден. Обратитесь к администратору.")

# Остальные клиенты
with st.expander("Пятерочка", expanded=False):
    st.markdown("""
    ### Инструкция по запуску

    **ВАЖНО!** Скрипт нужно запускать в Google Colab:

    1. Откройте [Google Colab](https://colab.research.google.com/)
    2. Создайте новую ячейку
    3. Скопируйте код ниже и вставьте в ячейку
    4. Запустите код (нажмите ▶️ или Shift+Enter)
    5. Загрузите файлы когда появится запрос

    #### Ожидаемые файлы:
    - **"Отчет История лидов ВР"** или **"История лидов"** - отчет из SKILLAZ
    - **"Raw_Data"** - данные из FINEBI

    #### Результат:
    - Файл **SKILLAZ_FINEBI_merged.xlsx** с двумя вкладками:
      - **FINEBI** - данные FINEBI с добавленными столбцами из SKILLAZ
      - **SKILLAZ** - исходные данные SKILLAZ
    """)

    st.markdown("---")

    # Код сверки
    st.markdown("### Код для Google Colab")
    st.info("**Кнопка копирования** в правом верхнем углу блока кода скопирует **весь код целиком**")

    pyaterochka_code = """# Google Colab скрипт для объединения данных SKILLAZ и FINEBI
# Скопируйте этот код в Google Colab и запустите

# Установка библиотек и импорты
!pip install pandas openpyxl -q

import pandas as pd
import numpy as np
from google.colab import files
import re
from io import BytesIO

# Загрузка файлов
print("Загрузите два файла:")
print("1. Файл с 'Отчет История лидов ВР' в названии (SKILLAZ)")
print("2. Файл с 'Raw_Data' в названии (FINEBI)")
print()

uploaded = files.upload()

skillaz_file = None
finebi_file = None

for filename in uploaded.keys():
    if 'Отчет История лидов' in filename or 'История лидов' in filename:
        skillaz_file = filename
        print(f"SKILLAZ файл: {filename}")
    elif 'Raw_Data' in filename:
        finebi_file = filename
        print(f"FINEBI файл: {filename}")

if not skillaz_file:
    print("\\n⚠️ Не найден файл SKILLAZ. Укажите вручную:")
    skillaz_file = list(uploaded.keys())[0]
    print(f"Используется: {skillaz_file}")

if not finebi_file:
    print("\\n⚠️ Не найден файл FINEBI. Укажите вручную:")
    finebi_file = list(uploaded.keys())[1] if len(uploaded) > 1 else list(uploaded.keys())[0]
    print(f"Используется: {finebi_file}")

# Загрузка и обработка данных SKILLAZ
df_skillaz = pd.read_excel(BytesIO(uploaded[skillaz_file]))

print(f"\\nSKILLAZ загружен: {len(df_skillaz)} строк")
print(f"Столбцы: {list(df_skillaz.columns)}")

# Обработка id_отклика - извлекаем первую часть до дефиса
if 'id отклика' in df_skillaz.columns:
    df_skillaz['id_отклика_clean'] = df_skillaz['id отклика'].astype(str).apply(
        lambda x: x.split('-')[0] if pd.notna(x) and x != 'nan' else ''
    )
elif 'id_отклика' in df_skillaz.columns:
    df_skillaz['id_отклика_clean'] = df_skillaz['id_отклика'].astype(str).apply(
        lambda x: x.split('-')[0] if pd.notna(x) and x != 'nan' else ''
    )
else:
    # Ищем столбец с похожим названием
    id_col = [c for c in df_skillaz.columns if 'отклик' in c.lower() and 'id' in c.lower()]
    if id_col:
        df_skillaz['id_отклика_clean'] = df_skillaz[id_col[0]].astype(str).apply(
            lambda x: x.split('-')[0] if pd.notna(x) and x != 'nan' else ''
        )
        print(f"Использован столбец: {id_col[0]}")

# Определяем столбец с датой статуса
date_col = None
for col in df_skillaz.columns:
    if 'дата статуса' in col.lower():
        date_col = col
        break

if date_col:
    # Преобразуем в datetime
    df_skillaz[date_col] = pd.to_datetime(df_skillaz[date_col], errors='coerce')
    # Сортируем по дате статуса по возрастанию
    df_skillaz = df_skillaz.sort_values(by=date_col, ascending=True)
    print(f"Отсортировано по столбцу: {date_col}")
else:
    print("⚠️ Столбец 'дата статуса' не найден")

print(f"\\nПример id_отклика после обработки: {df_skillaz['id_отклика_clean'].head().tolist()}")

# Загрузка и обработка данных FINEBI
df_finebi = pd.read_excel(BytesIO(uploaded[finebi_file]))

print(f"\\nFINEBI загружен: {len(df_finebi)} строк")
print(f"Столбцы: {list(df_finebi.columns)}")

# Находим столбец response_id (может быть с пробелом)
response_col = None
for col in df_finebi.columns:
    if 'response_id' in col.lower().replace(' ', '_'):
        response_col = col
        break

if response_col:
    # Перемещаем response_id в конец
    cols = [c for c in df_finebi.columns if c != response_col]
    cols.append(response_col)
    df_finebi = df_finebi[cols]

    # Приводим к строке для сопоставления
    df_finebi['response_id_str'] = df_finebi[response_col].astype(str).str.strip()
    print(f"Столбец response_id перемещен в конец")
else:
    print("⚠️ Столбец response_id не найден!")

# Сопоставление данных (аналог ПРОСМОТР)
# Определяем столбцы для переноса (без "дата статуса")
cols_to_transfer = []
for col in df_skillaz.columns:
    col_lower = col.lower()
    if 'skillaz статус' in col_lower or col_lower == 'skillaz статус':
        cols_to_transfer.append(col)
    elif 'vr статус' in col_lower:
        cols_to_transfer.append(col)
    elif col.lower() == 'этап':
        cols_to_transfer.append(col)

print(f"\\nСтолбцы для переноса: {cols_to_transfer}")

# Создаем словарь с последними значениями для каждого id_отклика
# Поскольку данные отсортированы по возрастанию даты, последнее значение - это последняя строка
last_values = {}

for idx, row in df_skillaz.iterrows():
    id_clean = str(row['id_отклика_clean']).strip()
    if id_clean and id_clean != 'nan' and id_clean != '':
        # Перезаписываем - так как данные отсортированы по возрастанию,
        # последняя запись будет самой свежей
        last_values[id_clean] = {col: row[col] for col in cols_to_transfer}

print(f"Уникальных id_отклика в SKILLAZ: {len(last_values)}")

# Добавление столбцов в FINEBI
for col in cols_to_transfer:
    new_col_name = col
    df_finebi[new_col_name] = df_finebi['response_id_str'].apply(
        lambda x: last_values.get(str(x).strip(), {}).get(col, 'Нет данных')
    )

# Удаляем вспомогательный столбец
df_finebi = df_finebi.drop(columns=['response_id_str'])

# Статистика сопоставления
matched = df_finebi[df_finebi[cols_to_transfer[0]] != 'Нет данных'].shape[0] if cols_to_transfer else 0
total = len(df_finebi)
print(f"\\nСопоставлено: {matched} из {total} строк ({matched/total*100:.1f}%)")

# Подготовка SKILLAZ для экспорта
# Удаляем вспомогательный столбец id_отклика_clean из SKILLAZ для экспорта
df_skillaz_export = df_skillaz.drop(columns=['id_отклика_clean'], errors='ignore')

print(f"\\nSKILLAZ подготовлен: {len(df_skillaz_export)} строк, {len(df_skillaz_export.columns)} столбцов")
print(f"FINEBI подготовлен: {len(df_finebi)} строк, {len(df_finebi.columns)} столбцов")

# Экспорт в Excel с двумя вкладками
output_filename = 'SKILLAZ_FINEBI_merged.xlsx'

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    df_finebi.to_excel(writer, sheet_name='FINEBI', index=False)
    df_skillaz_export.to_excel(writer, sheet_name='SKILLAZ', index=False)

print(f"\\n✅ Файл создан: {output_filename}")
print(f" - Вкладка FINEBI: {len(df_finebi)} строк")
print(f" - Вкладка SKILLAZ: {len(df_skillaz_export)} строк")

# Скачивание файла
files.download(output_filename)
print("\\n✅ Готово! Файл скачивается...")
"""

    with st.expander("Показать код", expanded=False):
        st.code(pyaterochka_code, language="python", line_numbers=False)


with st.expander("Магнит", expanded=False):
    st.info("Сверка для этого клиента находится в разработке")

# Закрываем контейнер matrix-code-section
st.markdown('</div>', unsafe_allow_html=True)

st.markdown("---")
st.markdown(
    "Сделано с ❤️ | Данные из API HH.ru",
    unsafe_allow_html=True
)
